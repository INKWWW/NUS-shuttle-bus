#!/usr/bin/python
# -*-coding: UTF-8-*-
# @Author :Wang Hanmo
# 
# Calculate the ratios of points in different fences and pick up the most highest one as the choosen sevice
# Use these points not in depots 

### import ###
from openpyxl import load_workbook   #pip install openpyxl
from collections import namedtuple
import csv
import myfunction
import math
import pdb

start_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'Car Park 11', 'Botanic Gardens MRT']
stop_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'Car Park 11', 'BTC - Oei Tiong Ham Building']
headers = ['gps_time', 'node_id','vehicle_serial','latitude', 'longitude', 'altitude', 'speed', 'heading', 'POI', 'start_stop', 'service']
service = ['A1', 'A2', 'B1', 'B2', 'C', 'D1', 'D2', 'A1E', 'A2E', 'BTC1', 'BTC2', 'UT-FoS'] #12
all_polygon = {}

### get all the vertexs on the NUS_fence orderly
wpoly = load_workbook('../NUS_polygon.xlsx')
sheet = wpoly.get_sheet_names()
poly = wpoly.get_sheet_by_name(sheet[0])
all_rows = poly.max_row
all_columns = poly.max_column
print('polygon_columns : ', all_columns)
print('polygon_rows : ', all_rows)
polygon = []
coordinate = ()
j = 0
for col in range (1,37,3):
    j += 1
    polygon = []
    for i in range(3, all_rows + 1) :
        coordinate = (poly.cell(row = i, column = col).value, poly.cell(row = i, column = col+1).value)
        if coordinate[0] is not None:
            polygon.append(coordinate)
    all_polygon[service[j-1]] = polygon
print(all_polygon)
### breakpoint
# pdb.set_trace()

### use geo fence algorithm to calculate the ratios of points in fences ###
counter_state = -1  # -1:negative   1:active
counter_all = 0
counter_in = 0
counter_row = 0
record_row_service_start = []
record_row_service_stop = []
times_service = 0
ratio_service = {}
ratio = []

for ser in service:
    counter_state = -1  # -1:negative   1:active
    counter_all = 0
    counter_in = 0
    counter_row = 0
    ratio = []
    with open('ss_vehicle/ss_2025_inside-2017-10-23.csv', 'r') as ss_v:  #with打开不用在意close
        r_ss = csv.reader(ss_v)
        headings_v = next(r_ss)
        # print(headings_v)              
        Vehicle = namedtuple('Vehicle', headings_v)   #用namedtuple 方便后面用headings来读取数据
        for row_v in r_ss:
            counter_row += 1
            # print(row_v[8])
            # print(row_v[8] == 'EA')
            # print(len(row_v) == 9)
            # pdb.set_trace()
            
            if (len(row_v) == 10) and (row_v[len(row_v)-1] == 'stop'):
                counter_state = -1
            if (len(row_v) == 10) and (row_v[len(row_v)-1] == 'start'):
                counter_state = 1

            if counter_state == 1:
                counter_all += 1
                ### Geo Fence
                in_or_out = myfunction.point_in_poly(float(row_v[3]), float(row_v[4]), all_polygon[ser])
                print(in_or_out)
                if in_or_out is 'OUT':
                    continue
                else :
                    counter_in += 1

            if (counter_state == -1) and (counter_in > 100) :   #counter_in is bigger than any value(maybe minimum 5) to prevent some small bias
                times_service += 1
                record_row_service_start.append(counter_row - counter_all)
                record_row_service_stop.append(counter_row)                
                ratio.append(counter_in / counter_all)
                ratio_service[ser] = ratio
                counter_all = 0
                counter_in = 0
print(ratio_service)
# print(record_row_service_start)
# print(record_row_service_stop)
## reverse the ratio_service
# new_ratio_service = {v:k for k,v in ratio_service.items()}
# print(new_ratio_service)

#### pick up the service according to the highest ratio
raw_ratio = {}
service_list = []
for i in range(0, times_service):
    for ser in service:
        raw_ratio[ser] = ratio_service[ser][i]
    new_raw_service = {v:k for k,v in service.items()}
    max_ratio = max(list(new_raw_service.keys()))
    service_list.append(new_raw_service[max_ratio])
print('service_list')
print(service_list)


with open('ss_vehicle/ss_2025_inside-2017-10-23.csv', 'r') as ss_v:  #with打开不用在意close
    r_ss = csv.reader(ss_v)
    headings_ss = next(r_ss)             
    Vehicle = namedtuple('Vehicle', headings_v)   #用namedtuple 方便后面用headings来读取数据
    headings_ss.pop()
    headings_ss.pop()
    print(headings_ss)             
    SS = namedtuple('SS', headings_ss)   #用namedtuple 方便后面用headings来读取数据

    with open('service_vehicle/service_2025_2017-10-23.csv', 'w', newline ='') as fnew:    #use 'w' for writing str and newline='' for deleting extra idle row
        w_s = csv.writer(fnew)
        w_s.writerow(headers)
        counter = 0
        for i in range(0, times_service):
            for row_s in r_ss:
                counter += counter

                if counter in range(record_row_service_start, record_row_service_stop+1):
                    row_s.append('')
                    row_s.append(service_list[i])
                    w_ss.writerow([row_s[0], row_s[1], row_s[2], row_s[3],row_s[4], row_s[5], row_s[6], \
                        row_s[7], row_s[8], row_s[10]])




            
            






