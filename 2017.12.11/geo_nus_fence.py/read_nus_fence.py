#!/usr/bin/python
# -*-coding: UTF-8-*-
# @Author :Wang Hanmo

from openpyxl import load_workbook   #pip install openpyxl

### load and read the  List of NUS Shuttle Bus POIs.xlsx
wpoly = load_workbook('../NUS_polygon.xlsx')
sheet = wpoly.get_sheet_names()
poly = wpoly.get_sheet_by_name(sheet[0])
all_rows = poly.max_row
all_columns = poly.max_column
print('polygon_columns : ', all_columns)
print('polygon_rows : ', all_rows)

polygon = []
coordinate = ()
for i in range(3, all_rows + 1) :
    coordinate = (poly.cell(row = i, column = 1).value, poly.cell(row = i, column = 2).value)
    polygon.append(coordinate)
print(polygon)




# poligono = [(-33.416032,-70.593016), (-33.415370,-70.589604),
# (-33.417340,-70.589046), (-33.417949,-70.592351),
# (-33.416032,-70.593016)]
# print(type(poligono))




