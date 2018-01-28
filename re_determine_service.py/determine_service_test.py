#!/usr/bin/python
# -*-coding: UTF-8-*-
# @Author :Wang Hanmo
# 
# Calculate the ratios of points in different fences and pick up the most highest one as the choosen sevice
# Use these points not in depots 

### import ###
from openpyxl import load_workbook   #pip install openpyxl
from collections import namedtuple
import csv
import myfunction
import math
import copy
import pdb  #breakpoint debug

start_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'Car Park 11', 'Botanic Gardens MRT']
stop_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'Car Park 11', 'BTC - Oei Tiong Ham Building']
headers = ['gps_time', 'node_id','vehicle_serial','latitude', 'longitude', 'altitude', 'speed', 'heading', 'POI', 'start_stop', 'service']
service = ['A1', 'A2', 'B1', 'B2', 'C', 'D1', 'D2', 'A1E', 'A2E', 'BTC1', 'BTC2', 'UT-FoS'] #12
all_polygon = {}

### get all the vertexs on the NUS_fence orderly ###
wpoly = load_workbook('../NUS_polygon.xlsx')
sheet = wpoly.get_sheet_names()
poly = wpoly.get_sheet_by_name(sheet[0])
all_rows = poly.max_row
all_columns = poly.max_column
print('max_polygon_columns : ', all_columns)
print('max_polygon_rows : ', all_rows)

## Get all the polygons' coordinates
polygon = []
coordinate = ()
j = 0
for col in range (4,39,3):
    j += 1
    polygon = []
    for i in range(3, all_rows + 1) :
        coordinate = (poly.cell(row = i, column = col).value, poly.cell(row = i, column = col+1).value)
        # print(coordinate)
        if coordinate[0] is not None:
            polygon.append(coordinate)
    all_polygon[service[j-1]] = polygon
print(all_polygon)
### breakpoint
# pdb.set_trace()
# 
### Get the stations in order of every service ###
poi_read = load_workbook('../verify_service_route2.xlsx')
poi_sheet = poi_read.get_sheet_names()
poi = poi_read.get_sheet_by_name(poi_sheet[0])
poi_all_row = poi.max_row
poi_all_column = poi.max_column
print('--------------------------POI:----------------------------')
print('max_poi_rows: ', poi_all_row)
print('max_poi_columns: ', poi_all_column)

print(poi.cell(row = 15, column = 1).value)
print(type(poi.cell(row = 15, column = 1).value))
# pdb.set_trace()

all_poi_normal = {}
poi_normal = []
poi_order = ''
all_poi_normal_num = {}
j = 0
for poi_col in range (1, 13):
    j += 1
    poi_normal = []
    for i in range (2, poi_all_row + 1):
        poi_order = poi.cell(row = i, column = poi_col).value
        # print(poi_order)
        if poi_order is not None:
            poi_normal.append(poi_order)
        else:
            continue
    all_poi_normal[service[j-1]] = poi_normal
    all_poi_normal_num[service[j-1]] = len(poi_normal)
print(all_poi_normal)
print(all_poi_normal_num)
# pdb.set_trace()

### use geo fence algorithm to calculate the ratios of points in fences ###
counter_state = -1  # -1:negative   1:active
counter_all = 0
counter_in = 0
counter_row = 0
record_row_service_start = []
record_row_service_stop = []
times_service = 0
ratio_service = {}
ratio = []
times_list = {}
service_count_list = []

for ser in service:
    counter_state = -1  # -1:negative   1:active
    counter_all = 0
    counter_in = 0
    counter_row = 0
    counter_state_begin = 0 # 0:negative   1:active
    times_service_begin = 0 # 0:negative   1:active
    times_service = 0
    ratio = []
    with open('ss_vehicle/ss_2033_inside-2017-09-04.csv', 'r') as ss_v:  #with打开不用在意close
        r_ss = csv.reader(ss_v)
        headings_v = next(r_ss)
        # print(headings_v)              
        # Vehicle = namedtuple('Vehicle', headings_v)   #用namedtuple 方便后面用headings来读取数据
        for row_v in r_ss:
            counter_row += 1
            # print(row_v[8])
            # print(row_v[8] == 'EA')
            # print(len(row_v) == 9)
            # pdb.set_trace()
            
            if (len(row_v) == 10) and (row_v[len(row_v)-1] == 'stop'):
                counter_state = -1
                
            if (len(row_v) == 10) and (row_v[len(row_v)-1] == 'start'):
                counter_state = 1

            if (len(row_v) == 9):
                counter_state_begin = 1                

            if (counter_state == 1) and (counter_state_begin == 1):
                times_service_begin = 1
                counter_all += 1

                ### Geo Fence
                in_or_out = myfunction.point_in_poly(float(row_v[3]), float(row_v[4]), all_polygon[ser])
                print(in_or_out)
                if in_or_out is 'OUT':
                    continue
                else :
                    counter_in += 1

            # if (counter_state == -1) and (counter_in > 200) :   #counter_in is bigger than any value(maybe minimum 5) to prevent some small bias
            if (counter_state == -1) and (counter_state_begin == 1) and (times_service_begin == 1)\
                and (counter_all > 360):    #limit the rows of counter_all to reduce 'false' routes -- 6min
                times_service += 1
                times_service_begin = 0
                counter_state_begin = 0
                record_row_service_start.append(counter_row - counter_all)
                record_row_service_stop.append(counter_row)                
                ratio.append(counter_in / counter_all)
                ratio_service[ser] = ratio
                counter_all = 0
                counter_in = 0
    times_list[ser] = times_service
print('-----------------------------------------------------------------------')
print('all ratios are: \n %s' %ratio_service)
print('-----------------------------------------------------------------------')
print('all the times is: \n %s' %times_list)
# pdb.set_trace()
print(record_row_service_start)
print(record_row_service_stop)
pdb.set_trace()
## reverse the ratio_service
# new_ratio_service = {v:k for k,v in ratio_service.items()}
# print(new_ratio_service)

### pick up the service according to the highest ratio ###
raw_ratio = {}
service_list = []
for i in range(0, times_service):
    for ser in service:
        raw_ratio[ser] = ratio_service[ser][i]
    print('-----------------------------------------------------------------------')
    print('raw_ratio: \n')
    print(raw_ratio)
    new_raw_ratio = sorted(raw_ratio.items(), key = lambda item : item[1], reverse = True)
    print('new_raw_ratio: \n')
    print(new_raw_ratio)
    service_list.append(new_raw_ratio[0][0])
    print('i is :---- ', (service.index(new_raw_ratio[0][0])+ i * 10))
print('-----------------------------------------------------------------------')
print('service_list:')
print(service_list)
# pdb.set_trace()

## use service's whole POIs to match service_list to substract unnormal routes
with open('ss_vehicle/ss_2033_inside-2017-09-04.csv', 'r') as ss_v:  #with打开不用在意close
    # r_ss = csv.reader(ss_v)
    counter_poi = 0
    pop_index = 0
    final_list = copy.deepcopy(service_list)
    for i in range(0, times_service):
        temp_poi = []
        # ss_v.seek(record_row_service_start[i], 0)
        poi_start = record_row_service_start[i]
        r_ss = csv.reader(ss_v)
        for poi_row in r_ss:
            # print(poi_row)
            counter_poi += 1
            if counter_poi >= poi_start:
                poi_start += 1
                if poi_start <= record_row_service_stop[i]: 
                    # print(poi_row[8])
                    # pdb.set_trace()                       
                    if poi_row[8] not in temp_poi:
                        temp_poi.append(poi_row[8])
                else:
                    print(temp_poi)
                    if len(temp_poi) >= all_poi_normal_num[service_list[i]]:
                        print("-----: ", service_list)
                        pop_index += 1
                        break
                    else:
                        final_list.pop(pop_index)
                        print('pop---: ', final_list)
                        break
print('Final_service_list: \n', final_list)
pdb.set_trace()

    # new_raw_service = {v:k for k,v in raw_ratio.items()}  #reverse the dictionary
    # print('-----------------------------------------------------------------------')
    # print('new_raw_service: \n')
    # print(new_raw_service)
    # max_ratio = max(list(new_raw_service.keys()))
    # service_list.append(new_raw_service[max_ratio])
# print('-----------------------------------------------------------------------')
# print('max_ratio:')
# print(max_ratio)

temp_final_service = ''
service_sort_set = set()
service_sort_set = sorted(list(set(service_list)), key=service_list.index)
print('-----------------------------------------------------------------------')
print('service_sort_set:')
print(service_sort_set)

service_count_list = [0] * len(service_sort_set)
for i in range(len(service_sort_set)):
    service_count_list[i] = service_list.count(service_sort_set[i])
temp_final_service = service_sort_set[service_count_list.index(max(service_count_list))]
print('-----------------------------------------------------------------------')
print('temp_final_service : ')
print(temp_final_service)

## differentiate A series bus
A1_state = 0  #0:negative  1:active
A2_state = 0  #0:negative  1:active
with open('ss_vehicle/ss_2033_inside-2017-09-04.csv', 'r') as ss_v:  #with打开不用在意close
    r_ss = csv.reader(ss_v)
    for row_s in r_ss:

        ## A
        if ( (temp_final_service == 'A1') or (temp_final_service == 'A2') or (temp_final_service == 'A1E') \
            or (temp_final_service == 'A2E') ):

            final_service = 'A1E'
            
            #1. A2
            if row_s[8] == 'Museum':
                
                final_service = 'A2'
                A2_state = 1
                break
            #2. A1
            if row_s[8] == 'COM 2 (CP13)':
                final_service = 'A1'
                A1_state = 1
                break
        ## B
        elif ( (temp_final_service == 'B1') or (temp_final_service == 'B2') ):
            final_service = 'B1'
            if row_s[8] == 'EA':
                final_service = 'B2'
                break

        else:
            final_service = temp_final_service
    print('final_service: %s' %final_service)
pdb.set_trace()

    
### write these files with the final)service ###
with open('ss_vehicle/ss_2033_inside-2017-09-04.csv', 'r') as ss_v:  #with打开不用在意close
    r_ss = csv.reader(ss_v)

    with open('service_vehicle/service_2033_2017-09-04.csv', 'w', newline ='') as fnew:    #use 'w' for writing str and newline='' for deleting extra idle row
        w_s = csv.writer(fnew)
        # w_s.writerow(headers)
        counter = 0
        for row_s in r_ss:
            counter += 1
            if (len(row_s) == 10):
                w_s.writerow([row_s[0], row_s[1], row_s[2], row_s[3],row_s[4], row_s[5], row_s[6], \
                        row_s[7], row_s[8], row_s[9], final_service])
            else:
                w_s.writerow([row_s[0], row_s[1], row_s[2], row_s[3],row_s[4], row_s[5], row_s[6], \
                        row_s[7], row_s[8], ' ', final_service])
print('success!')            



                


