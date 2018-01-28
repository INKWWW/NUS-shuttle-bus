#!/usr/bin/python
# -*-coding: UTF-8-*-
# @Author :Wang Hanmo
# 
### import ###
from openpyxl import load_workbook   #pip install openpyxl
from collections import namedtuple
import csv
import myfunction
import math
import copy
import pprint
import pdb  #breakpoint debug

start_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'BIZ2', 'Botanic Gardens MRT', 'Ventus (Opp LT13)', 'University Town']
# stop_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'Car Park 11', 'BTC - Oei Tiong Ham Building']
headers = ['gps_time', 'node_id','vehicle_serial','latitude', 'longitude', 'altitude', 'speed', 'heading', 'POI', 'start_stop', 'service']
service = ['A1', 'A2', 'B1', 'B2', 'C*', 'C', 'D1', 'D2', 'A1E', 'A2E', 'BTC1', 'BTC2', 'UT-FoS'] #13
start_POI = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'BIZ2', 'Opp HSSML', 'Botanic Gardens MRT', 'University Town']
must_stop_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'Botanic Gardens MRT']
all_polygon = {}
read_file_name = 'fenced_vehicle/2051_inside-2017-09-04.csv'

### get all the vertexs on the NUS_fence orderly ###
def get_NUS_fence_vertex():
    wpoly = load_workbook('../NUS_polygon.xlsx')
    sheet = wpoly.get_sheet_names()
    poly = wpoly.get_sheet_by_name(sheet[0])
    all_rows = poly.max_row
    all_columns = poly.max_column
    print('max_polygon_columns : ', all_columns)
    print('max_polygon_rows : ', all_rows)
## Get all the polygons' coordinates
    polygon = []
    coordinate = ()
    j = 0
    for col in range (4,39,3):
        j += 1
        polygon = []
        for i in range(3, all_rows + 1) :
            coordinate = (poly.cell(row = i, column = col).value, poly.cell(row = i, column = col+1).value)
            # print(coordinate)
            if coordinate[0] is not None:
                polygon.append(coordinate)
        all_polygon[service[j-1]] = polygon
    print(all_polygon)
### breakpoint
# pdb.set_trace()

### Get the stations in order of every service ###
def getStations_Inorder():
    poi_read = load_workbook('../verify_service_route5.xlsx')
    poi_sheet = poi_read.get_sheet_names()
    poi = poi_read.get_sheet_by_name(poi_sheet[0])
    poi_all_row = poi.max_row
    poi_all_column = poi.max_column
    print('--------------------------POI:----------------------------')
    print('max_poi_rows: ', poi_all_row)
    print('max_poi_columns: ', poi_all_column)
    # print(poi.cell(row = 15, column = 1).value)
    # print(type(poi.cell(row = 15, column = 1).value))
    # pdb.set_trace()
    all_poi_normal = {}
    poi_normal = []
    poi_order = ''
    all_poi_normal_num = {}
    j = 0
    for poi_col in range (1, 14):
        j += 1
        poi_normal = []
        for i in range (2, poi_all_row + 1):
            poi_order = poi.cell(row = i, column = poi_col).value
            # print(poi_order)
            if poi_order is not None:
                poi_normal.append(poi_order)
            else:
                continue
        all_poi_normal[service[j-1]] = poi_normal
        all_poi_normal_num[service[j-1]] = len(poi_normal)
    pprint.pprint(all_poi_normal)
    # print(all_poi_normal_num)
    return(all_poi_normal)
    # print('------')
    # print(all_poi_normal.items())
    # pdb.set_trace()

### match the service according to the order of POIs in different service ###
def match_Service(all_poi_normal, all_poi_normal_num):
    
    service_copy = copy.deepcopy(service)
    start_POI_copy = copy.deepcopy(start_POI)
    temp_service_list = []
    matching_service = []
    A1_A1E_dis = []
    C_CKV_dis =[]
    LT29_counter = 0
    zero_speed_counter = 0
    counter = 0
    mid_appear_counter = 0
    start_match = 0
    counter_start_mark = 0
    A1E_match_counter = 0
    must_stop_counter = 0
    biz2_stop_counter = 0
    LT29_counter_state = True
    # particular_must_stop_poi = []
    match_start_state = True
    check = False

    pointer_State = {}     # record_start:True
    for i in service_copy:
        pointer_State[i] = True
    # print(pointer_State)

    pointer_POI = {}
    for i in service_copy:
        pointer_POI[i] = -1
    print(pointer_POI)
    # pdb.set_trace()

    with open(read_file_name, 'r') as ss_v:  #with打开不用在意close
        r_ss = csv.reader(ss_v)
        for row in r_ss:
            counter += 1
            # print(counter)
            ## define the start of the process of matching
            
            # Because there are only two A2E in the afternoon (about 17:50)
            if match_start_state and row[8] == 'Ventus (Opp LT13)':
                A1E_match_counter += 1
                if A1E_match_counter > 15:                    
                    time_hour = int(row[0].split('T')[1].split(':')[0])
                    time_min = int(row[0].split('T')[1].split(':')[1])
                    if time_hour >= 17:
                        pointer_POI['A2E'] = 0
                        match_start_state = False
                        A1E_match_counter = 0
                        matching_service = [key for key, values in pointer_POI.items() if values == 0]
                        if len(matching_service) > 0:
                            print('@@ ', matching_service)
                        print('counter:  ', counter)
                        print('time: ', row[0])
            else:
                A1E_match_counter = 0

            # others except A2E
            if match_start_state and row[8] in start_POI_copy:
                counter_start_mark += 1
            else:
                counter_start_mark = 0

            if match_start_state and counter_start_mark > 0:               
                for service_item in service_copy:              
                    if (all_poi_normal[service_item][0] == row[8]) and (row[8] in start_POI_copy):
                        pointer_POI[service_item] = 0
                        # pointer_State[service_item] = False
                    if service_item == 'UT-FoS' and (0 in [values for key, values in pointer_POI.items()]):
                        match_start_state = False
                matching_service = [key for key, values in pointer_POI.items() if values == 0]
                if len(matching_service) > 0:
                    print('@@ ', matching_service)
                print('counter: ----- ', counter)
                print('time:', row[0])            

            #D2  VS  UT-FoS  // D2 pass 2 times of LT29
            if row[8] != 'LT29':
                LT29_counter_state = True
            if row[8] == 'LT29' and LT29_counter_state:
                LT29_counter += 1
                LT29_counter_state = False

            ## match
            for i in matching_service:

                # continuously, velocity = 0 for a long time --> re_match
                if int(row[6]) == 0 and row[8] != all_poi_normal[i][0]:
                    zero_speed_counter += 1
                if int(row[6]) != 0:
                    zero_speed_counter = 0

                # A1  VS  A1E
                if ('A1' in matching_service) and ('A1E' in matching_service):
                    if row[8] not in A1_A1E_dis or row[8] != A1_A1E_dis[len(A1_A1E_dis) - 1]:
                        A1_A1E_dis.append(row[8])

                if row[8] == all_poi_normal[i][pointer_POI[i]]:
                    # print(row)
                    pointer_POI[i] += 1
                    print(pointer_POI)
                    if pointer_POI[i] == 2:
                        print('start_service_time: ', row[0])

                    # # A1  VS  A1E
                    # if ('A1' in matching_service) and ('A1E' in matching_service):
                    #     if row[8] not in A1_A1E_dis:
                    #         A1_A1E_dis.append(row[8])

                    # C  VS  C*
                    if ('C' in matching_service) and ('C*' in matching_service):
                        if row[8] not in C_CKV_dis:
                            C_CKV_dis.append(row[8])

                        # if row[8] not in A1_A1E_dis:
                        #     A1_A1E_dis.append(row[8])
                            # print('A1_A1E_dis: ', A1_A1E_dis)
                    # pdb.set_trace()
                    
                if (row[8] != all_poi_normal[i][0]):    ##!!!!!improvement
                    check = True
                    # print('!!!',counter)
                
                # prevent the ahead apperance of first station  --> quit and restart
                if pointer_POI[i] > 0 and check:
                    if row[8] == all_poi_normal[i][0]:
                        mid_appear_counter += 1   #在for循环里面，每行都会自加len(matching_service)次
                        # print(row)
                        # print('-?-', mid_appear_counter)
                    else:
                        mid_appear_counter = 0
                        
                    # must stop and restart matching
                    if row[8] in must_stop_point and ('BTC1' not in matching_service):
                        must_stop_counter += 1    #在for循环里面，每行都会自加len(matching_service)次
                        # print('must stop poi:', must_stop_counter)
                        # print('counter is ', counter)
                    if ('B2' in matching_service or 'D1' in matching_service) and row[8] == 'BIZ2' and int(row[6]) <= 1:
                        biz2_stop_counter += 1
                        # print('biz2_stop_counter: ',biz2_stop_counter)
                    if row[8] != 'BIZ2':
                        biz2_stop_counter = 0

                    if (mid_appear_counter // len(matching_service) >= 35) or \
                        (must_stop_counter // len(matching_service) >= 35) or \
                        (biz2_stop_counter // len(matching_service) >= 50) or \
                        (zero_speed_counter // len(matching_service) >= 300):
                        match_start_state = True
                        check = False
                        mid_appear_counter = 0
                        must_stop_counter = 0
                        biz2_stop_counter = 0
                        counter_start_mark = 0
                        zero_speed_counter = 0
                        LT29_counter = 0
                        matching_service = []
                        A1_A1E_dis = []
                        print('pause counter: -- ', counter)
                        # if pointer_POI[i] < all_poi_normal_num[i]:
                        #     start_POI_copy.pop(start_POI_copy.index(all_poi_normal[i][0]))
                        for j in service_copy:
                            # pointer_State[j] = True
                            pointer_POI[j] = -1
                        break                       

                # determine the final service    --> quit and restart
                if pointer_POI[i] == all_poi_normal_num[i]:
                    # A1  VS A1E
                    # if (i == 'A1E') and ('COM 2 (CP13)' in A1_A1E_dis):
                    print('A1_A1E_dis: ', A1_A1E_dis)
                    if i == 'A1E' and A1_A1E_dis[A1_A1E_dis.index('COM 2 (CP13)') + 1] == 'Opp NUSS':
                        # print('A1_A1E_dis: ', A1_A1E_dis)
                        matching_service.pop(matching_service.index('A1E'))
                        # print('pop :', matching_service)
                        A1_A1E_dis = []
                        break

                    # D2  VS  UT-FoS
                    if (i == 'D2') and LT29_counter > 3:
                        i = 'UT-FoS'
                        # print('LT29_counter:', LT29_counter)

                    # C  VS  C*
                    if i == 'C*' and 'Computer Centre' in C_CKV_dis:
                        matching_service.pop(matching_service.index('C*'))
                        C_CKV_dis = []
                        break                    

                    temp_service_list.append(i)
                    print('SUCCESS----: ', i)
                    print('success_counter:  ', counter)
                    matching_service = []
                    match_start_state = True
                    mid_appear_counter = 0
                    must_stop_counter = 0
                    biz2_stop_counter = 0
                    counter_start_mark = 0
                    LT29_counter = 0
                    A1_A1E_dis = []
                    # start_POI_copy = copy.deepcopy(start_POI)
                    for j in service_copy:
                        # pointer_State[j] = True
                        pointer_POI[j] = -1
                    break
                        
    print('result: \n', temp_service_list)
    print('length: ', len(temp_service_list))




############## main ###############
if __name__ == "__main__":
    # print('---------------------------------- main -----------------------------------')
    all_poi_normal_num = {}
    all_poi_normal = getStations_Inorder()
    # print(all_poi_normal)
    # pdb.set_trace()
    for i in service:
        all_poi_normal_num[i] = len(all_poi_normal[i])

    print('all_poi_normal_num: ', all_poi_normal_num)

    match_Service(all_poi_normal, all_poi_normal_num)








