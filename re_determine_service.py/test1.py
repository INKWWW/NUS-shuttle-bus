#!/usr/bin/python
# -*-coding: UTF-8-*-
# @Author :Wang Hanmo

from openpyxl import load_workbook
from collections import namedtuple
import csv
import myfunction

# ### get all the vertexs on the NUS_fence orderly
# wpoly = load_workbook('../NUS_polygon.xlsx')
# sheet = wpoly.get_sheet_names()
# poly = wpoly.get_sheet_by_name(sheet[0])
# all_rows = poly.max_row
# all_columns = poly.max_column
# print('polygon_columns : ', all_columns)
# print('polygon_rows : ', all_rows)
# polygon = []
# coordinate = ()
# for i in range(3, all_rows + 1) :
#     coordinate = (poly.cell(row = i, column = 1).value, poly.cell(row = i, column = 2).value)
#     polygon.append(coordinate)
# print(polygon)

# in_or_out = myfunction.point_in_poly(1.29315, 103.78539, polygon)
# print(in_or_out)

# a = 'ok'
# if (type(a) is str) and (a is 'ok' or a == 1 or a == 2):
#     print('true')
# else:
#     print('false')

# b = '3'
# print(type(int(b)))

# a = {}
# a['A1'] = [1,2]
# a['B2'] = 2
# # print(a.values())
# # order = list(a.values())
# # new_a = {v:k for k,v in a.items()}
# # print(new_a)
# # print(new_a[order[0]])
# print(a['A1'][0])
# break

# sss = ''
# ccc = ['a', 'b', 'c']
# sss = '#'.join(ccc)
# print(sss)

# n = 2  
# m = 3  
# k = 2
#   
# matrix = [ [ [None]*2 for w in range(3) ] for e in range(2) ]
# print(matrix)

# for i in range(len(matrix)):  
#     matrix[i] = [0]*3  
# print(matrix)
  
# for i in range(n):  
#     for j in range(m):  
#         matrix[i][j] = [1]*k
# print(matrix)  

# matrix[0].append([2,2])
# print(matrix)

# matrix = [ [ [None]*4 for w in range(3) ] for e in range(2) ]
# print(matrix)

# LIST = [0] * 3
# print(LIST)

ll = [1,1,1,2,3,4,5,6,2,3,1,2,3,6,5]
print(ll.index(max(ll)))

