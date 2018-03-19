#!/usr/bin/python
# -*-coding: UTF-8-*-
# @Author :Wang Hanmo

#################################################Geo Fencing############################################
'''1. Clean up data and rearrange data. This module is designed to seperate particular bus from the raw data
    downloaded from DataCommon that will be used in Module 'DetermineService.py' to detect service.
   2. The Geo Fence's Polygon is NUS Area(Kent Ridge Campus & Botanic Campus), used to eliminate the location data outside NUS campus
    limit the coverage inside NUS Campus Area. In addition, tag the locations of buses
    
    First Step
'''

from openpyxl import load_workbook   #pip install openpyxl
from collections import namedtuple
import csv
import MyFunctions
import math
import pdb


def ListOfPoi(filepath):
    '''Load and Read the "List of NUS Shuttle Bus POIs.xlsx"
    '''
    wpoi = load_workbook(filepath)
    sheet = wpoi.get_sheet_names()
    poi = wpoi.get_sheet_by_name(sheet[0])
    return poi

def NusPolygon(filepath):
    '''Get all the vertexs on the NUS_fence orderly
    '''
    wpoly = load_workbook(filepath)
    sheet = wpoly.get_sheet_names()
    poly = wpoly.get_sheet_by_name(sheet[0])
    all_rows = poly.max_row - 1
    all_columns = poly.max_column
    print('polygon_columns : ', all_columns)
    print('polygon_rows : ', all_rows)
    polygon_NUS = []
    coordinate = ()
    for i in range(3, all_rows + 1) :
        coordinate = (poly.cell(row = i, column = 1).value, poly.cell(row = i, column = 2).value)
        polygon_NUS.append(coordinate)
    # print(polygon_NUS)
    return polygon_NUS

def AllBusNodeid(filepath):
    '''Extract all vehicles running in a particular day - Get all node_id
    '''
    with open(filepath, 'r') as f_vehicle:  #with打开不用在意close
        r_vehicle = csv.reader(f_vehicle)
        headings_v = next(r_vehicle)
        # print(headings_v)              
        Vehicle = namedtuple('Vehicle', headings_v)   #用namedtuple 方便后面用headings来读取数据
        nodeid_list = []
        nodeid_license = {}
        nodeid_counter = 0
        for row_v in r_vehicle:
            vehicle = Vehicle._make(row_v)
            if vehicle.node_id not in nodeid_list:
                nodeid_list.append(vehicle.node_id)
                nodeid_counter = nodeid_counter + 1
                nodeid_license[vehicle.node_id] = vehicle.vehicle_serial
        print('all node_id today: ', nodeid_counter)
        print('node_id_list: ', nodeid_list)
        print('nodeid_license: ', nodeid_license)
        return nodeid_list


def GeoFence_NusPoly(filepath, poi, polygon_NUS, nodeid_list):
    '''Geo Fence for each particular bus
       Clean up points outside the NUS_Polygon

       filepath: file 
    '''   
    headers = ['node_id', 'vehicle_serial', 'gps_time', 'latitude', 'longitude', 'altitude', 'speed', 'heading', 'POI', 'service']
    Depot_zone = ['Car Park 11 A', 'Car Park 11 B', 'Car Park 11 C', 'Car Park 11 D', 'Kent Ridge Terminal A', \
              'Kent Ridge Terminal B', 'Kent Ridge Terminal C', 'Kent Ridge Terminal D','PGP Terminal A', \
              'PGP Terminal B', 'PGP Terminal C', 'PGP Terminal D']

    for vehicleid in nodeid_list:
        with open(filepath, 'r') as f_vehicle:  #with打开不用在意close
            r_vehicle = csv.reader(f_vehicle)
            headings_v = next(r_vehicle)
            # print(headings_v)              
            Vehicle = namedtuple('Vehicle', headings_v)   #用namedtuple 方便后面用headings来读取数据
            newfilename = '-'.join([vehicleid, filepath.split('/')[-1]])  #根据不同的车辆号码输入创建不同的文件名
            newpath = '/'.join(['fenced_vehicle', filepath.split('/')[-1].split('.')[0], newfilename])
            print(newpath)
            MyFunctions.checkPath(newpath)
            with open(newpath, 'w', newline ='') as fnew:    #use 'w' for writing str and newline='' for deleting extra idle row
                w_vehicle = csv.writer(fnew)
                w_vehicle.writerow(headers)
                for row_v in r_vehicle:
                    vehicle = Vehicle._make(row_v)
                    # print(vehicle.node_id + '111')
                    if vehicle.node_id == vehicleid:

                        # Geo Fencing inside NUS campus
                        in_or_out = MyFunctions.point_in_poly(float(vehicle.latitude), float(vehicle.longitude), polygon_NUS)
                        print(in_or_out)
                        if in_or_out is 'OUT':
                            continue
                        else :  
                            # add 8 to the original time -- Singapore Zone
                            time = vehicle.gps_time
                            date_time = int(time.split('T')[0].split('-')[2])
                            hour_time = int(time.split('T')[1].split(':')[0])
                            minsec_time = time.split(':')[1] + ':' + time.split(':')[2]
                            if (hour_time + 2) >= 24:
                                date_time = date_time + 1
                            else:
                                date_time = date_time
                            if (hour_time + 8) > 24:
                                hour_time = hour_time + 8 - 24
                            else:
                                hour_time = hour_time + 8
                            new_time = time.split('T')[0].split('-')[0] + '-' + time.split('T')[0].split('-')[1] + '-' + str(date_time) + 'T' \
                                + ' ' + str(hour_time) + ':' + minsec_time       

                            # calculate the distance from the information about latitude and longitude
                            poi_rows = list(range(2, 48))
                            distance = {}                
                            for poi_row in poi_rows:
                                poi_latitude = poi.cell(row = poi_row, column = 2).value
                                poi_longitude = poi.cell(row = poi_row, column = 3).value
                                vehicle_latitude = float(vehicle.latitude)
                                vehicle_longitude = float(vehicle.longitude)                
                                distance.update({poi.cell(row = poi_row, column = 1).value : \
                                    MyFunctions.cal_distance(poi_latitude, poi_longitude, vehicle_latitude, vehicle_longitude)})
                            
                            ### sort the distance
                            distance_sort_values = sorted(distance.values())                
                            smallest_distance = distance_sort_values[0]
                            #print(smallest_distance)
                            new_distance = {v:k for k,v in distance.items()}  #reverse the original key:value to find the nearest point
                            nearest_point = new_distance[smallest_distance]
                            #print(nearest_point)

                            ### calculate the slope for deciding whether vihicle is in Depot or not
                            cp_A_la = poi.cell(row = 37, column = 2).value; cp_A_lo = poi.cell(row = 37, column = 3).value
                            cp_B_la = poi.cell(row = 38, column = 2).value; cp_B_lo = poi.cell(row = 38, column = 3).value
                            cp_C_la = poi.cell(row = 39, column = 2).value; cp_C_lo = poi.cell(row = 39, column = 3).value
                            cp_D_la = poi.cell(row = 40, column = 2).value; cp_D_lo = poi.cell(row = 40, column = 3).value
                            krt_A_la = poi.cell(row = 41, column = 2).value; krt_A_lo = poi.cell(row = 41, column = 3).value
                            krt_B_la = poi.cell(row = 42, column = 2).value; krt_B_lo = poi.cell(row = 42, column = 3).value
                            krt_C_la = poi.cell(row = 43, column = 2).value; krt_C_lo = poi.cell(row = 43, column = 3).value
                            krt_D_la = poi.cell(row = 44, column = 2).value; krt_D_lo = poi.cell(row = 44, column = 3).value
                            pgp_A_la = poi.cell(row = 45, column = 2).value; pgp_A_lo = poi.cell(row = 45, column = 3).value
                            pgp_B_la = poi.cell(row = 46, column = 2).value; pgp_B_lo = poi.cell(row = 46, column = 3).value
                            pgp_C_la = poi.cell(row = 47, column = 2).value; pgp_C_lo = poi.cell(row = 47, column = 3).value
                            pgp_D_la = poi.cell(row = 48, column = 2).value; pgp_D_lo = poi.cell(row = 48, column = 3).value
                            slope_cp = MyFunctions.cal_slope(cp_A_la, cp_A_lo, cp_D_la, cp_D_lo)
                            slope_krt = MyFunctions.cal_slope(krt_A_la, krt_A_lo, krt_D_la, krt_D_lo)
                            slope_pgp = MyFunctions.cal_slope(pgp_A_la, pgp_A_lo, pgp_C_la, pgp_C_lo)

                            #decide the POI
                            #Campus_zone = ['NUS Kent Ridge (Centroid)', 'UTown Centre']
                            Depot_zone = ['Car Park 11 A', 'Car Park 11 B', 'Car Park 11 C', 'Car Park 11 D',\
                            'Kent Ridge Terminal A', 'Kent Ridge Terminal B', 'Kent Ridge Terminal C', 'Kent Ridge Terminal D',\
                            'PGP Terminal A', 'PGP Terminal B', 'PGP Terminal C', 'PGP Terminal D']
                            
                            Right_place = ''
                            start_stop = ''
                            if nearest_point in Depot_zone:
                                index = Depot_zone.index(nearest_point)
                                #calculate the slope 
                                if index < 4:  # Depot: Car park 11
                                    slope_A = MyFunctions.cal_slope(cp_A_la, cp_A_lo, vehicle_latitude, vehicle_longitude)
                                    slope_B = MyFunctions.cal_slope(cp_B_la, cp_B_lo, vehicle_latitude, vehicle_longitude)
                                    slope_C = MyFunctions.cal_slope(cp_C_la, cp_C_lo, vehicle_latitude, vehicle_longitude)
                                    slope_D = MyFunctions.cal_slope(cp_D_la, cp_D_lo, vehicle_latitude, vehicle_longitude)
                                    if slope_A < 0:
                                        slope_A = slope_A + math.pi
                                    if slope_B < 0:
                                        slope_B = slope_B + math.pi
                                    if slope_C < 0:
                                        slope_C = slope_C + math.pi
                                    if slope_D < 0:
                                        slope_D = slope_D + math.pi

                                    if slope_A == 0 or slope_B == 0 or slope_C == 0 or slope_D == 0:
                                        Right_place = 'Car Park 11'
                                    
                                    elif math.atan(slope_A) <= math.atan(slope_cp) and math.atan(slope_B) >= math.atan(slope_cp) \
                                        and math.atan(slope_C) <= math.atan(slope_cp) and math.atan(slope_D) >=  math.atan(slope_cp):
                                        Right_place = 'Car Park 11'
                                        
                                    else:
                                        n = 1
                                        next_distance = distance_sort_values[n]
                                        Right_place = new_distance[next_distance]
                                        while( Right_place in Depot_zone):
                                            n = n + 1
                                            next_distance = distance_sort_values[n]
                                            Right_place = new_distance[next_distance]


                                elif index > 3 and index < 8:   #Depot: Kent Ridge Terminal
                                    slope_A = MyFunctions.cal_slope(krt_A_la, krt_A_lo, vehicle_latitude, vehicle_longitude)
                                    slope_B = MyFunctions.cal_slope(krt_B_la, krt_B_lo, vehicle_latitude, vehicle_longitude)
                                    slope_C = MyFunctions.cal_slope(krt_C_la, krt_C_lo, vehicle_latitude, vehicle_longitude)
                                    slope_D = MyFunctions.cal_slope(krt_D_la, krt_D_lo, vehicle_latitude, vehicle_longitude)
                                    if slope_A < 0:
                                        slope_A = slope_A + math.pi
                                    if slope_B < 0:
                                        slope_B = slope_B + math.pi
                                    if slope_C < 0:
                                        slope_C = slope_C + math.pi
                                    if slope_D < 0:
                                        slope_D = slope_D + math.pi

                                    if slope_A == 0 or slope_B == 0 or slope_C == 0 or slope_D == 0:
                                        Right_place = 'Kent Ridge Bus Terminal'
                                        
                                    elif math.atan(slope_A) >= math.atan(slope_krt) and math.atan(slope_B) <= math.atan(slope_krt) \
                                        and math.atan(slope_C) >= math.atan(slope_krt) and math.atan(slope_D) <=  math.atan(slope_krt):
                                        Right_place = 'Kent Ridge Bus Terminal'
                                        
                                    else:
                                        n = 1
                                        next_distance = distance_sort_values[n]
                                        Right_place = new_distance[next_distance]
                                        while( Right_place in Depot_zone):
                                            n = n + 1
                                            next_distance = distance_sort_values[n]
                                            Right_place = new_distance[next_distance]
                                else:  #Depot: Prince George's Park Terminal
                                    slope_A = MyFunctions.cal_slope(pgp_A_la, pgp_A_lo, vehicle_latitude, vehicle_longitude)
                                    slope_B = MyFunctions.cal_slope(pgp_B_la, pgp_B_lo, vehicle_latitude, vehicle_longitude)
                                    slope_C = MyFunctions.cal_slope(pgp_C_la, pgp_C_lo, vehicle_latitude, vehicle_longitude)
                                    slope_D = MyFunctions.cal_slope(pgp_D_la, pgp_D_lo, vehicle_latitude, vehicle_longitude)
                                    if slope_A < 0:
                                        slope_A = slope_A + math.pi
                                    if slope_B < 0:
                                        slope_B = slope_B + math.pi
                                    if slope_C < 0:
                                        slope_C = slope_C + math.pi
                                    if slope_D < 0:
                                        slope_D = slope_D + math.pi

                                    if slope_A == 0 or slope_B == 0 or slope_C == 0 or slope_D == 0:
                                        Right_place = 'Prince George\'s Park Terminal'
                                        
                                    elif math.atan(slope_A) >= math.atan(slope_pgp) and math.atan(slope_B) >= math.atan(slope_pgp) \
                                        and math.atan(slope_C) <= math.atan(slope_pgp) and math.atan(slope_D) <=  math.atan(slope_pgp):
                                        Right_place = 'Prince George\'s Park Terminal'
                                        
                                    else:
                                        n = 1
                                        next_distance = distance_sort_values[n]
                                        Right_place = new_distance[next_distance]
                                        while( Right_place in Depot_zone):
                                            n = n + 1
                                            next_distance = distance_sort_values[n]
                                            Right_place = new_distance[next_distance]
                            else:
                                Right_place = nearest_point;

                            w_vehicle.writerow([row_v[0], row_v[1], new_time, row_v[3],row_v[4], row_v[5], row_v[6], row_v[7], Right_place])
                   

def run(filepath):
    '''aggregate all these functions and run this module
        
        filepath: Raw Data's path for a particular day 
    '''
    poi = ListOfPoi('../New List of NUS Shuttle Bus POIs.xlsx')
    polygon_NUS = NusPolygon('../NUS_polygon.xlsx')
    nodeid_list = AllBusNodeid(filepath)
    GeoFence_NusPoly(filepath, poi, polygon_NUS, nodeid_list)
    print('Success!')


if __name__ == '__main__':
    filepath = '../Veniam_BusLocation/2018_week02/2018-01-09.csv'
    run(filepath)




















