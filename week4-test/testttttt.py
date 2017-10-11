#!/usr/bin/python
# -*-coding: UTF-8-*-
# @Author :Wang Hanmo

from openpyxl import load_workbook   #pip install openpyxl
from collections import namedtuple
import csv
import math
import myfunctions

wpoi = load_workbook('New List of NUS Shuttle Bus POIs.xlsx')
sheet = wpoi.get_sheet_names()
poi = wpoi.get_sheet_by_name(sheet[0])
print(poi.title)
# print(poi.cell(row=12, column=3).value)
# print(type(poi.cell(row=12, column=3).value) )
cp_A_la = poi.cell(row = 37, column = 2).value; cp_A_lo = poi.cell(row = 37, column = 3).value
cp_D_la = poi.cell(row = 40, column = 2).value
cp_D_lo = poi.cell(row = 40, column = 3).value
slope_cp = myfunctions.cal_slope(cp_A_la, cp_A_lo, cp_D_la, cp_D_lo)
print(slope_cp)
print(math.atan(-1) * 180 / math.pi)
print(1  0)



# rows = poi.rows
# n = 0
# for row in rows:
#     n += 1
# print(n)  

# c = {'a':3, 'b':1, 'c':2}
# point = c.keys()
# print(point)
# cc = sorted(c.values(), reverse = True)
# print(cc)
# print(cc[0])
# new_c = {v:k for k,v in c.items()}
# print(new_c[cc[1]])

# print(int('2'))


# with open('../Veniam_BusLocation/2017_week35/2017-08-28.csv', 'r') as f_vehicle:  #with打开不用在意close
#     r_vehicle = csv.reader(f_vehicle)
#     headings_v = next(r_vehicle)
#     # print(headings_v)              
#     Vehicle = namedtuple('Vehicle', headings_v)   #用namedtuple 方便后面用headings来读取数据
#     for row_v in r_vehicle:
#         vehicle = Vehicle._make(row_v)
#         latitude = vehicle.latitude
#         print(type(latitude))
#         break


# result = myfunctions.cal_distance(1.298985, 103.774179, 1.306537, 103.772861)
# print(result)
# aa = list(range(1,3))
# print(aa)