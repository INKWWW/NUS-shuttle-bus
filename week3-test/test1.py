#!/usr/bin/python
# -*-coding: UTF-8-*-
# @Author :Wang Hanmo

#determine the vehicle's start of service
from openpyxl import load_workbook   #pip install openpyxl
from collections import namedtuple
import csv
import calculate_distance

#load and read the  List of NUS Shuttle Bus POIs.xlsx
wpoi = load_workbook('../New List of NUS Shuttle Bus POIs.xlsx')
sheet = wpoi.get_sheet_names()
poi = wpoi.get_sheet_by_name(sheet[0])

#require to input the targeted vehicle_id
vehicleid = input('please input the vehicleid: ')  
headers = ['gps_time', 'node_id','vehicle_serial','latitude', 'longitude', 'POI','altitude', 'speed', 'heading']
counter_stop = 0

#read the data from vehicle and create a file to save data
with open('../Veniam_BusLocation/2017_week35/2017-08-28.csv', 'r') as f_vehicle:  #with打开不用在意close
    r_vehicle = csv.reader(f_vehicle)
    headings_v = next(r_vehicle)
    # print(headings_v)              
    Vehicle = namedtuple('Vehicle', headings_v)   #用namedtuple 方便后面用headings来读取数据
    newfilename = '_'.join([vehicleid, 'test-2017-08-28.csv'])  #根据不同的车辆号码输入创建不同的文件名
    with open(newfilename, 'w', newline ='') as fnew:    #use 'w' for writing str and newline='' for deleting extra idle row
        w_vehicle = csv.writer(fnew)
        w_vehicle.writerow(headers)
        for row_v in r_vehicle:
            vehicle = Vehicle._make(row_v)
            # print(vehicle.node_id + '111')
            if vehicle.node_id == vehicleid:
                #add 8 to the original time -- Singapore Zone
                time = vehicle.gps_time
                date_time = int(time.split('T')[0].split('-')[2])
                hour_time = int(time.split('T')[1].split(':')[0])
                minsec_time = time.split(':')[1] + ':' + time.split(':')[2]
                if (date_time + 1) == 28:
                    date_time = date_time + 1
                else:
                    date_time = date_time
                if (hour_time + 8) > 24:
                    hour_time = hour_time + 8 - 24
                else:
                    hour_time = hour_time + 8
                new_time = time.split('T')[0].split('-')[0] + '-' + time.split('T')[0].split('-')[1] + '-' + str(date_time) + 'T' \
                    + ' ' + str(hour_time) + ':' + minsec_time
                
                #calculate the distance from the information anout latitude and longitude
                poi_rows = list(range(2, 47))
                distance = {}                
                for poi_row in poi_rows:
                    poi_latitude = poi.cell(row = poi_row, column = 2).value
                    poi_longitude = poi.cell(row = poi_row, column = 3).value
                    vehicle_latitude = float(vehicle.latitude)
                    vehicle_longitude = float(vehicle.longitude)                
                    distance.update({poi.cell(row = poi_row, column = 1).value : \
                        calculate_distance.cal_dis(poi_latitude, poi_longitude, vehicle_latitude, vehicle_longitude)})

                
                #sort the distance
                distance_sort_values = sorted(distance.values())                
                smallest_distance = distance_sort_values[0]
                #print(smallest_distance)
                new_distance = {v:k for k,v in distance.items()}  #reverse the original key:value to find the nearest point
                nearest_point = new_distance[smallest_distance]
                #print(nearest_point)

                #decide the POI
                Campus_zone = ['NUS Kent Ridge (Centroid)', 'UTown Centre']
                Depot_zone = ['Car Park 11 zone 1', 'Car Park 11 zone 2', \
                'Car Park 11 zone 3', 'Car Park 11 zone 4',\
                'Car Park 11 zone 5', 'Kent Ridge Terminal zone 1', \
                'Kent Ridge Terminal zone 2', 'Kent Ridge Terminal zone 3',\
                'PGP Terminal zone 1']
                #whether in Campus Zone
                Right_place = ''
                if nearest_point in Campus_zone:
                    index = Campus_zone.index(nearest_point)
                    radius = poi.cell(row = 37 + index, column = 3).value
                    if smallest_distance <= radius:
                        Right_place = nearest_point
                        #print(nearest_point + ' is in Campus zone')
                    else:
                        #find the second nearest point
                        second_distance = distance_sort_values[1]                       
                        second_point = new_distance[second_distance]
                        Right_place = second_point


                elif nearest_point in Depot_zone:
                    index = Depot_zone.index(nearest_point)
                    radius = poi.cell(row = 39 + index, column = 3).value
                    if smallest_distance <= radius:
                        Right_place = nearest_point
                        #print(nearest_point + ' is in Depot zone')
                    else:
                        #find the second nearest point
                        second_distance = distance_sort_values[1]                       
                        second_point = new_distance[second_distance]
                        Right_place = second_point

                else:
                    Right_place = nearest_point

                w_vehicle.writerow([new_time, row_v[0],row_v[1],row_v[3],row_v[4], Right_place, row_v[5],row_v[6],row_v[7]])
               