#!/usr/bin/python
# -*-coding: UTF-8-*-
# @Author :Wang Hanmo
# 
'''Detect service and tag them'''
from openpyxl import load_workbook   #pip install openpyxl
from collections import namedtuple
import pandas as pd
import csv
import MyFunctions
import math
import copy
import pprint
import os
import pdb  #breakpoint debug

start_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'BIZ2', 'Botanic Gardens MRT', 'Ventus (Opp LT13)', 'University Town']
# stop_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'Car Park 11', 'BTC - Oei Tiong Ham Building']
headers = ['node_id','vehicle_serial','gps_time','latitude', 'longitude', 'altitude', 'speed', 'heading', 'POI', 'service', 'service_start_stop']  # 11
service = ['A1', 'A2', 'B1', 'B2', 'C*', 'C', 'D1', 'D2', 'A1E', 'A2E', 'BTC1', 'BTC2', 'UT-FoS'] #13
start_POI = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'BIZ2', 'Opp HSSML', 'Botanic Gardens MRT', 'University Town']
must_stop_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'Botanic Gardens MRT']
all_polygon = {}
# 用于判断一天中间数据缺失，而且是在服务一开始最初选择candidate的阶段。若final_service_list已有这些站点中一个，
# 但是matching_list里面没有，则用一个list去存经过的站点，再做概率计算，推测可能性
indepent_service = ['A1', 'A2', 'B1', 'B2', 'C*', 'C', 'D1', 'D2']

# read_file_name = '../fenced_vehicle/2024_inside_2017-09-04.csv'

'''Visit all files in a folder and get these files' names'''
def VisitAllFile(folderpath):
    fileName = [filenames for (dirpath, dirnames, filenames) in os.walk(folderpath)]
    print('FILENAME: ', fileName)
    return fileName

'''Get all the vertexs on the NUS_fence orderly'''
def get_NUS_fence_vertex():
    wpoly = load_workbook('../NUS_polygon.xlsx')
    sheet = wpoly.get_sheet_names()
    poly = wpoly.get_sheet_by_name(sheet[0])
    all_rows = poly.max_row
    all_columns = poly.max_column
    # print('max_polygon_columns : ', all_columns)
    # print('max_polygon_rows : ', all_rows)
    polygon = []
    coordinate = ()
    j = 0
    for col in range (4,39,3):
        j += 1
        polygon = []
        for i in range(3, all_rows + 1) :
            coordinate = (poly.cell(row = i, column = col).value, poly.cell(row = i, column = col+1).value)
            # print(coordinate)
            if coordinate[0] is not None:
                polygon.append(coordinate)
        all_polygon[service[j-1]] = polygon
    # print(all_polygon)
### breakpoint
# pdb.set_trace()

'''Get the stations in order of every service'''
def getStations_Inorder():
    poi_read = load_workbook('../verify_service_route_7.xlsx')
    poi_sheet = poi_read.get_sheet_names()
    poi = poi_read.get_sheet_by_name(poi_sheet[0])
    poi_all_row = poi.max_row
    poi_all_column = poi.max_column
    # print('--------------------------POI:----------------------------')
    # print('max_poi_rows: ', poi_all_row)
    # print('max_poi_columns: ', poi_all_column)
    # print(poi.cell(row = 15, column = 1).value)
    # print(type(poi.cell(row = 15, column = 1).value))
    # pdb.set_trace()
    all_poi_normal = {}
    poi_normal = []
    poi_order = ''
    all_poi_normal_num = {}
    all_POI = []
    j = 0
    for poi_col in range (1, 14):
        j += 1
        poi_normal = []
        for i in range (2, poi_all_row + 1):
            poi_order = poi.cell(row = i, column = poi_col).value
            # print(poi_order)
            if poi_order is not None:
                poi_normal.append(poi_order)
            else:
                continue
        all_poi_normal[service[j-1]] = poi_normal   ##{'A1':[XX, XX, XX, XX], 'B2':[XX, XX, XX, XX]...}
        all_poi_normal_num[service[j-1]] = len(poi_normal)
    # pprint.pprint(all_poi_normal)
    # print(all_poi_normal_num)
    # 
    ## 统计所有的站点POI
    for key in all_poi_normal.keys():
        for m in all_poi_normal[key]:
            if m not in all_POI:
                all_POI.append(m)
    # print('all_POI:', all_POI)
    # print(len(all_POI))    

    return(all_poi_normal)  
    # print('------')
    # print(all_poi_normal.items())
    # pdb.set_trace()

'''Match the Service according to the order of POIs in different service'''
def match_Service(filepath, all_poi_normal, all_poi_normal_num):
    
    service_copy = copy.deepcopy(service)
    start_POI_copy = copy.deepcopy(start_POI)
    final_service_list = []
    matching_service = []
    A1_A1E_dis = []
    C_CKV_dis =[]
    time = []
    final_time = []
    start_row = []
    stop_row = []
    service_time_pair = []
    start_service_time = {}
    start_row_1 = {}
    start_row_2 = []
    LT29_counter = 0
    zero_speed_counter = 0
    counter = 0
    mid_appear_counter = 0
    start_match = 0
    counter_start_mark = 0
    A2E_match_counter = 0
    must_stop_counter = 0
    biz2_stop_counter = 0
    LT29_counter_state = True
    # particular_must_stop_poi = []
    match_start_state = True
    A2E_match_state = True
    check = False  #用于检测车辆在可能服务期间返回起始站（即提前出现起始站点），作为检验的准入条件控制
    start_service_time_state = True

    # 针对UT-FoS的判断
    UT_FoS_hour_list = [9, 11, 13, 15]
    UT_FoS_match_counter = 0
    UTFOS_time_hour = 0
    UT_FoS_match_state = False

    # 针对一天中最开始的第一个服务缺失的情况，在所有服务中计算概率，所得概率值最大值对应的服务,则算是(service-P)-'service-predict'
    # 注意：每天的第一趟服务的所有站点都会被记录，因为假设是缺失的，但是只有在第一趟潜在服务被判断成pause的时候才计算该概率
    miss_headmost_passby_poi = []   #记录经过的站点，知道pause的时候停止记录，开始计算概率
    miss_headmost_prob = {}   #{A1:prob_x, B1:prob_y, ....}   prob_x为概率值
    miss_headmost_prob_self = {}
    miss_headmost_state = True   #每个service的第一趟服务才置为True
    first_row_time = ''  #记录第一行的时间

    # 针对一天中除第一个service之外的中间的service缺失的情况，则将上一次的service-即final_service_list目前的最后一个service
    # 拿出来，作为最可能出现的下一个service,然后进行剩下站点的匹配，计算相应的概率
    # 随时都记录经过的站点，遇到完整的匹配或者pause的时候，所记录的站点清除
    miss_mid_passby_poi = []  #记录经过的站点，直到ause的时候停止记录，开始计算概率
    miss_mid_prob = {}  #目前只需要求上一个可能站点在这次缺失路线中的可能性，但是初始化为dict，作为扩展
    miss_mid_time_state = True  #计数使用，每次预匹配的时候记录一次
    miss_mid_start_row = 0
    miss_mid_start_time = ''
    
    pointer_State = {}     # record_start:True
    for i in service_copy:
        pointer_State[i] = True
    # print(pointer_State)

    pointer_POI = {}  #{'A1':-1, 'B2':-1 .....}
    for i in service_copy:
        pointer_POI[i] = -1
    # print(pointer_POI)
    # pdb.set_trace()

    with open(filepath, 'r') as ss_v:  #with打开不用在意close
        r_ss = csv.reader(ss_v)
        for row in r_ss:
            counter += 1
            # print(counter)
            ## Define the start of the process of matching
            
            ## 针对最早的一班service,开始记录每天第一趟服务的所经站点。遇到pause或者判断成功都要将miss_headmost_state置为False。遇到pause才计算概率。
            if counter == 2:
                first_row_time = row[2]               
            if row[8] not in miss_headmost_passby_poi and miss_headmost_state:
                miss_headmost_passby_poi.append(row[8])
                # print('miss_headmost_passby_poi', miss_headmost_passby_poi)
            
            ## 针对中途丢失数据的
            if row[8] not in miss_mid_passby_poi and len(final_service_list) != 0:
                miss_mid_passby_poi.append(row[8])
                # 此处需要使用try & except来防止第一次服务的时候final_service_list还为空
                try:
                    # small trick:当出现第一个可能是该服务的站点的时候才记录时间-->[:-1]
                    if miss_mid_time_state and row[8] in (all_poi_normal[final_service_list[-1]])[:-1]:
                        miss_mid_start_time = row[2]
                        miss_mid_start_row = counter
                        miss_mid_time_state = False
                except (IndexError, KeyError) as e:  #处理某service的最后一个round，因为最后一个round没有duration了！！
                    print(e)
                    print('Opps. But it does not matter! I can record them too!')
                    miss_mid_start_time = row[2]
                    miss_mid_start_row = counter
                    miss_mid_time_state = False

            ## A2E
            # Because there are only two A2E in the afternoon (about 17:30 & 17:40) - different vehicles runs these two service 
            # what's more, 不会让跑A1 或者 A2的车去跑
            if row[8] == 'Ventus (Opp LT13)' and A2E_match_state:
                A2E_match_counter += 1
                if A2E_match_counter > 12:                    
                    time_hour = int(row[2].split('T')[1].split(':')[0].strip())  #strip()消除字符前后可能出现的空格
                    time_min = int(row[2].split('T')[1].split(':')[1].strip())
                    if time_hour == 17 and ('A1' not in matching_service) and ('A2' not in matching_service):
                        pointer_POI['A2E'] = 0
                        A2E_match_counter = 0
                        matching_service.append('A2E')
                        start_service_time['A2E'] = row[2]
                        start_row_1['A2E'] = counter
                        # if len(matching_service) > 0:
                        #     print('@@@ A2E @@@ ', matching_service)
                        # print('counter:  ', counter)
                        # print('time: ', row[0])
                        # time.append(row[2])
            else:
                A2E_match_counter = 0

            # 决定UT-FoS是否需要加入到matching_list里面
            if (len(final_service_list) > 0) and ('D2' not in matching_service) and ('D2' not in final_service_list):
                UT_FoS_match_state = True
            if row[8] == 'University Town' and int(row[6]) == 0 and UT_FoS_match_state:
                UT_FoS_match_counter += 1
                if UT_FoS_match_counter >= 10:
                    UTFOS_time_hour = int(row[2].split('T')[1].split(':')[0].strip())  #strip()消除字符前后可能出现的空格
                    if UTFOS_time_hour in UT_FoS_hour_list:
                        pointer_POI['UT-FoS'] = 0
                        UT_FoS_match_counter = 0
                        matching_service.append('UT-FoS')
                        UT_FoS_hour_list.pop(UT_FoS_hour_list.index(UTFOS_time_hour))
                        start_service_time['UT-FoS'] = row[2]
                        start_row_1['UT-FoS'] = counter
                        UT_FoS_match_state = False
                        # if len(matching_service) > 0:
                        #     print('@@@ UT-FoS @@@ ', matching_service)
                        #     print('#$@#@%',UT_FoS_hour_list)               

            #D2  VS  UT-FoS  // D2 pass 2 times of LT29
            if row[8] != 'LT29':
                LT29_counter_state = True
            if row[8] == 'LT29' and LT29_counter_state:
                LT29_counter += 1
                LT29_counter_state = False

            # others except A2E
            if match_start_state and row[8] in start_POI_copy:
                counter_start_mark += 1
            else:
                counter_start_mark = 0

            # 开始匹配，并找出candidate
            if match_start_state and counter_start_mark > 0:             
                for service_item in service_copy:              
                    if (all_poi_normal[service_item][0] == row[8]) and (row[8] in start_POI_copy) and service_item != 'UT-FoS':
                        pointer_POI[service_item] = 0
                        #########
                        # start_service_time[service_item] = row[2]
                        # start_row_1[service_item] = counter
                        # pointer_State[service_item] = False
                    # 匹配到最后一个'UT-FoS'为止
                    if service_item == 'UT-FoS' and (0 in [values for key, values in pointer_POI.items()]):
                        match_start_state = False
                matching_service = [key for key, values in pointer_POI.items() if values == 0]  #such as [A1, A2, B1]
                # if len(matching_service) > 0:
                #     print('@@', matching_service)
                # print('counter: ----- ', counter)
            
            # C  VS  C*
            if ('C' in matching_service) and ('C*' in matching_service):
                if row[8] not in C_CKV_dis:
                    C_CKV_dis.append(row[8])

            # A1  VS  A1E
            if ('A1' in matching_service) or ('A1E' in matching_service):
                if row[8] not in A1_A1E_dis or row[8] != A1_A1E_dis[len(A1_A1E_dis) - 1]:
                    A1_A1E_dis.append(row[8])
        
            
            ###此处也要判断第一个预测服务的，针对不会遇到pause停止条件的情况！！！
            ## 针对每天第一趟服务的缺失进行判断 ##
            # if len(matching_service) != 0 and miss_headmost_state:
            #     if len(miss_headmost_passby_poi) != 0:
            #         print('PREDICT THE FIRST SERVICE')
                                            
            #         for miss_headmost_service in service:  #遍历所有服务来求概率
            #             miss_headmost_counter = 0
            #             for m in all_poi_normal[miss_headmost_service]:
            #                 if m in miss_headmost_passby_poi:
            #                     miss_headmost_counter += 1
            #                 else:
            #                     continue
            #             # float
            #             miss_headmost_prob[miss_headmost_service] = miss_headmost_counter / len(miss_headmost_passby_poi)
            #             miss_headmost_prob_self[miss_headmost_service] = miss_headmost_counter / all_poi_normal_num[miss_headmost_service]
            #         # print('miss_headmost_prob', miss_headmost_prob)
            #         # 遍历完成后，进行比较，选出maximum对应的服务
            #         miss_headmost_prob_reverse = {v:k for k,v in miss_headmost_prob.items()}  #反转键值对
            #         max_prob = max(miss_headmost_prob_reverse.keys())
            #         miss_headmost_predict = miss_headmost_prob_reverse[max_prob]
            #         if miss_headmost_state and miss_headmost_prob_self[miss_headmost_predict] >= float(0.5):
            #             # print('miss_headmost_predict', miss_headmost_predict)
            #             pointer_POI[miss_headmost_predict] = all_poi_normal_num[miss_headmost_predict]
            #             start_service_time[miss_headmost_predict] = first_row_time
            #             start_row.append(100)  #让后面计数的时候减100到0
            #             stop_row.append(counter)

            #         miss_headmost_passby_poi = []
            #         miss_headmost_state = False     
            #         # print('pointer_POI ---- :', pointer_POI)
            #         # print('matching_service', matching_service)
            #         final_service_list.append(miss_headmost_predict)
            #         final_time.append(first_row_time)      
                
                      

            # #D2  VS  UT-FoS  // D2 pass 2 times of LT29
            # if row[8] != 'LT29':
            #     LT29_counter_state = True
            # if row[8] == 'LT29' and LT29_counter_state:
            #     LT29_counter += 1
            #     LT29_counter_state = False


            ############################### match ##############################
            for i in matching_service:

                # # continuously, velocity = 0 for a long time --> re_match  移到下面pause条件后了
                # if int(row[6]) <= 2 and row[8] != all_poi_normal[i][0]:
                #     zero_speed_counter += 1
                # if int(row[6]) > 2:
                #     zero_speed_counter = 0
                # if i == 'D2' and int(row[2].split('T')[1].split(':')[0].strip()) == 8 and int(row[6]) <= 2:
                #     zero_speed_counter -= 1 

                # # A1  VS  A1E  移到上面去了,已有matching_service就可以做这个工作了
                # if ('A1' in matching_service) or ('A1E' in matching_service):
                #     if row[8] not in A1_A1E_dis or row[8] != A1_A1E_dis[len(A1_A1E_dis) - 1]:
                #         A1_A1E_dis.append(row[8])

                ## 记录时间  row[8] == all_poi_normal[i][0]-->很重要，确保是在匹配第一个站的时候来确定起始时间
                if pointer_POI[i] == 1 and int(row[6]) == 0 and row[8] == all_poi_normal[i][0] and \
                    start_service_time_state and (i != 'UT-FoS'):
                    start_service_time[i] = row[2]
                    start_row_1[i] = counter
                    # print("!!!!!start !!!!!!!", start_service_time)
                    # print('pointer_POI: ', pointer_POI)                    

                ## Match service from the first POI in a fixed order
                # 针对某些服务三个连续站点中,中间站点缺失的情况作出预测判断
                miss_mid_poi = [['LT13', 'AS7', 'BIZ2'], ['Opp HSSML', 'Opp NUSS', 'Ventus (Opp LT13)'], ['Museum', 'YIH', 'Central Library'],\
                                 ['Opp University Health Centre', 'YIH', 'Central Library'], ['LT13', 'AS7', 'COM 2 (CP13)']]
                miss_mid_match_state = False
                try:
                    for miss_cont in range(len(miss_mid_poi)):                        
                        if all_poi_normal[i][pointer_POI[i]-1] == miss_mid_poi[miss_cont][0]:
                            if all_poi_normal[i][pointer_POI[i]+1] == miss_mid_poi[miss_cont][2]:
                                miss_mid_match_state = True
                except IndexError as e:
                    print('all_poi_normal[i][pointer_POI[i]+1] exceed the Index of list')

                ############################ match 匹配的过程 ############################
                #第一行：针对普通正常无缺失的匹配
                #第二行：针对PGP Hse No.处的缺失进行匹配
                #第三行：针对UHall附近的缺失进行匹配
                #第四行：针对BIZ2附近的缺失
                PGP_Hse_list = ['PGP Hse No. 7', 'PGP Hse No. 12', 'PGP Hse No. 14 and 15', 'Opp PGP Hse No. 12']
                # print('--pointer_POI--', pointer_POI)
                if row[8] == all_poi_normal[i][pointer_POI[i]] or \
                    ((row[8] in PGP_Hse_list) and (all_poi_normal[i][pointer_POI[i]] in PGP_Hse_list)) or \
                    ((row[8] == 'Opp UHall') and (all_poi_normal[i][pointer_POI[i]] == 'UHall')) or \
                    ((row[8] == 'Car Park 11 (BIZ)') and (all_poi_normal[i][pointer_POI[i]] == 'BIZ2')) or \
                    miss_mid_match_state :
                    # print(row)
                    pointer_POI[i] += 1
                    miss_mid_match_state = False
                    
                    # 1、针对A1E，从A1中判断出A1E，因为A1E第一个站为A1的第二个站，所以需要把这个时间记录下来。在后面通过A1判断成
                    # A1E的时候作为A1E的起始时间
                    if pointer_POI[i] == 2:
                        # print('start_service_time: ', row[2])
                        # print('start_row : ', counter)
                        # print(start_service_time)
                        # print('pointer_POI: ', pointer_POI)
                        if row[2] not in time:
                            time.append(row[2])
                            start_row_2.append(counter)
                    
                    # A1  VS  A1E
                    # if ('A1' in matching_service) and ('A1E' in matching_service):
                    #     if row[8] not in A1_A1E_dis:
                    #         A1_A1E_dis.append(row[8])

                    # # C  VS  C*   移到上面去了。有matching_service之后就可以判断了 
                    # if ('C' in matching_service) and ('C*' in matching_service):
                    #     if row[8] not in C_CKV_dis:
                    #         C_CKV_dis.append(row[8])                  
                

                # 用于检测提前出现起始站点，如果行驶了一会儿后回到起始站，说明没有进行服务。需要在预判服务后（即要离开起始站后）才开始时才置True，然后进行判断。                
                # prevent the ahead apperance of first station  --> quit and restart
                if (row[8] != all_poi_normal[i][0]):    ##!!!!!improvement
                    check = True

                # 检验部分pause的条件
                if pointer_POI[i] > 0 and check:
                # if (1 in [values for key, values in pointer_POI.items()]) and check:
                    # 检验预测服务过程中，服务的起始站点提前出现并且速度为零-->终止匹配过程
                    if row[8] == all_poi_normal[i][0] and int(row[6]) == 0:
                        mid_appear_counter += 1   #在for循环里面，每行都会自加len(matching_service)次
                        # print(row)
                        # print('-?-', mid_appear_counter)
                    else:
                        mid_appear_counter = 0
                        
                    # must stop and restart matching
                    if row[8] in must_stop_point and ('BTC1' not in matching_service) and ('D2' not in matching_service):
                        must_stop_counter += 1    #在for循环里面，每行都会自加len(matching_service)次

                    # continuously, velocity = 0 for a long time --> re_match
                    if int(row[6]) <= 2 and row[8] != all_poi_normal[i][0]:
                        zero_speed_counter += 1
                    if int(row[6]) > 2:
                        zero_speed_counter = 0
                    if i == 'D2' and int(row[2].split('T')[1].split(':')[0].strip()) == 8 and int(row[6]) <= 2:
                        zero_speed_counter -= 1 
                    
                    # if ('B2' in matching_service or 'D1' in matching_service) and row[8] == 'BIZ2' and int(row[6]) <= 1:
                    #     biz2_stop_counter += 1
                        # print('biz2_stop_counter: ',biz2_stop_counter)
                    # if row[8] != 'BIZ2':
                    #     biz2_stop_counter = 0

                    ##### Pause for some reasons #####
                    if (mid_appear_counter // len(matching_service) >= 35) or \
                        (must_stop_counter // len(matching_service) >= 30) or \
                        (zero_speed_counter // len(matching_service) > 140):
                        # print('mid_appear_counter -- :', mid_appear_counter // len(matching_service))
                        # print('must_stop_counter -- :', must_stop_counter // len(matching_service))
                        # print('zero_speed_counter -- :', zero_speed_counter // len(matching_service))
                        # print(miss_mid_passby_poi)

                        match_start_state = True
                        check = False
                        start_service_time_state = True
                        mid_appear_counter = 0
                        must_stop_counter = 0
                        biz2_stop_counter = 0
                        counter_start_mark = 0
                        zero_speed_counter = 0
                        LT29_counter = 0
                        A2E_match_counter = 0
                        matching_service = []
                        A1_A1E_dis = []
                        start_service_time = {}
                        # UT_FoS_match_state = True
                        UT_FoS_match_counter = 0

                        time = []
                        
                        ## @@@ PAUSE @@@ ##
                        # print('pause counter: -- ', counter)

                        ### 在遇到pause的情况下-针对中间丢失数据的情况,用final_service_list里面的上一个站点，也就是当前记录的最后一个来计算概率(>0.6)
                        ## 70% 匹配率即可通过
                        if len(final_service_list) != 0:
                            miss_mid_counter = 0  #记录出现的符合的站点数
                            miss_mid_predict = final_service_list[-1]
                            miss_mid_poi_normal_copy = copy.deepcopy(all_poi_normal)
                            for miss_mid_poi_item in miss_mid_passby_poi:
                                try:
                                    if miss_mid_poi_item in miss_mid_poi_normal_copy[miss_mid_predict]:
                                        miss_mid_counter += 1
                                        miss_mid_poi_normal_copy[miss_mid_predict] = miss_mid_poi_normal_copy[miss_mid_predict][miss_mid_poi_normal_copy[miss_mid_predict].index(miss_mid_poi_item)+1:]
                                    else:
                                        continue
                                except KeyError as e:
                                    if e == 'UT':
                                        miss_mid_predict = 'UT-FoS'
                                    else:
                                        miss_mid_predict = final_service_list[-1].split('-')[0]
                                        miss_mid_counter += 1                                    
                            try:
                                miss_mid_prob[miss_mid_predict] = miss_mid_counter / all_poi_normal_num[miss_mid_predict]                        
                            except KeyError as e:
                                miss_mid_predict = 'UT-FoS'
                                miss_mid_prob[miss_mid_predict] = miss_mid_counter / all_poi_normal_num[miss_mid_predict]

                            miss_mid_counter = 0
                            if miss_mid_prob[miss_mid_predict] >= float(0.7):
                                # pointer_POI[miss_mid_predict] = all_poi_normal_num[miss_mid_predict]
                                final_time.append(miss_mid_start_time)
                                if miss_mid_start_row < stop_row[-1] + 60:
                                    start_row.append(stop_row[-1] + 123)  # 保证开始的行数正确
                                else:
                                    start_row.append(miss_mid_start_row + 60)  #让后面计数的时候减到正确的地方                                
                                stop_row.append(counter)
                                miss_mid_time_state = True  #让下次可以记录起始时间
                                final_service_list.append(miss_mid_predict + '-LDM' + '-' + str(miss_mid_prob[miss_mid_predict]))
                        miss_mid_passby_poi = []
                        miss_mid_time_state = True  #让下次可以记录起始时间
                        

                        ### 此处与上面那处不同，针对遇到pause停止条件的情况
                        ## 针对每天第一趟服务的缺失进行判断 ##
                        if len(miss_headmost_passby_poi) != 0 and miss_headmost_state:                                                   
                            for miss_headmost_service in service:  #遍历所有服务来求概率
                                miss_headmost_counter = 0
                                for m in all_poi_normal[miss_headmost_service]:
                                    if m in miss_headmost_passby_poi:
                                        miss_headmost_counter += 1
                                    else:
                                        continue
                                # float  应该用检测到的服务除以该服务应该经过的站点
                                # miss_headmost_prob[miss_headmost_service] = miss_headmost_counter / len(miss_headmost_passby_poi)
                                miss_headmost_prob_self[miss_headmost_service] = miss_headmost_counter / all_poi_normal_num[miss_headmost_service]
                            # 遍历完成后，进行比较，选出maximum对应的服务
                            # miss_headmost_prob_reverse = {v:k for k,v in miss_headmost_prob.items()}  #反转键值对
                            miss_headmost_prob_self_reverse = {v:k for k,v in miss_headmost_prob_self.items()}  #反转键值对
                            # max_prob = max(miss_headmost_prob_reverse.keys())
                            max_self_prob = max(miss_headmost_prob_self_reverse.keys())
                            # miss_headmost_predict = miss_headmost_prob_reverse[max_prob]
                            miss_headmost_predict = miss_headmost_prob_self_reverse[max_self_prob]
                            if miss_headmost_state and miss_headmost_prob_self[miss_headmost_predict] >= float(0.6):                                
                                # pointer_POI[miss_headmost_predict] = all_poi_normal_num[miss_headmost_predict]
                                start_service_time[miss_headmost_predict] = first_row_time
                                start_row.append(60)  #让后面计数的时候减100到0
                                stop_row.append(counter)
                                final_time.append(start_service_time[miss_headmost_predict])
                                final_service_list.append(miss_headmost_predict + '-LDH' + '-' + str(max_self_prob))
                            miss_headmost_passby_poi = []                        
                        ##### @@ 此处判断开头缺失值情况中,目前是否为第一趟服务，是的话则进入point_POI[service_predict]置为匹配完全-pointer_POI[i] == all_poi_normal_num[i]，否则break跳出 @@ #####
                        
                            miss_headmost_state = False
                            print('This is the first service in this day!!!')                        
                        
                        for j in service_copy:
                            # pointer_State[j] = True
                            pointer_POI[j] = -1
                        break


                ###### Find!! #####
                ## Have found the final service!! Then quit, reset all parameters and restart this program loop
                if pointer_POI[i] == all_poi_normal_num[i]:
                    # A1  VS A1E
                    # if (i == 'A1E') and ('COM 2 (CP13)' in A1_A1E_dis):
                    #     print('A1_A1E_dis: ', A1_A1E_dis)
                    # print(matching_service)

                    # the vehicle running 'A1' won't run the 'A1E' service   priority: or < and < not
                    if i == 'A1E' or i == 'A1' and 'A1' not in final_service_list:
                        if A1_A1E_dis.count('COM 2 (CP13)') == 1:
                            i = 'A1E'
                        if A1_A1E_dis.count('COM 2 (CP13)') == 2:
                            i = 'A1'

                    # if i == 'A1E' and A1_A1E_dis[A1_A1E_dis.index('COM 2 (CP13)') + 1] == 'Opp NUSS':
                    #     # print('A1_A1E_dis: ', A1_A1E_dis)
                    #     # matching_service.pop(matching_service.index('A1E'))
                    #     # print('pop :', matching_service)
                    #     i = 'A1'
                    #     print('A1E change to A1')
                    #     A1_A1E_dis = []
                    #     # break

                    # try:
                    #     if i == 'A1' and A1_A1E_dis[A1_A1E_dis.index('COM 2 (CP13)') + 1] == 'Opp HSSML':
                    #         i = 'A1E'
                    #         print('A1 change to A1E')
                    #         A1_A1E_dis = []
                    # except ValueError as e:
                    #     i = 'A1'
                    #     A1_A1E_dis = []


                    # D2  VS  UT-FoS
                    # scenario 1: 假如被判断成UT-FoS, 但是不在应该运行的时间内，则判断成'D2-S‘--特殊的D2，可能是确实了一些数据造成的误判
                    UT_FoS_hour = 0  
                    if (i == 'UT-FoS') and ('D2' in matching_service) and ('D2' in final_service_list):
                        UT_FoS_hour = int(row[2].split('T')[1].split(':')[0].strip())
                        if UT_FoS_hour not in UT_FoS_hour_list:
                            i = 'D2'
                            print('modify UT_FoS TO D2')

                    # scenatio 2: 假如被判断成D2，则判断经过了几个LT29，以此来判断是否为UT-FoS
                    if LT29_counter > 3 and ('D2' not in final_service_list):
                        UT_FoS_hour = int(row[2].split('T')[1].split(':')[0].strip())
                        if UT_FoS_hour in UT_FoS_hour_list:
                            i = 'UT-FoS'
                            print('modify D2 TP UT_FoS')
                        # print('LT29_counter:', LT29_counter)

                    # C  VS  C*
                    if i == 'C*' and 'Computer Centre' in C_CKV_dis:
                        matching_service.pop(matching_service.index('C*'))
                        C_CKV_dis = []
                        break

                    # Let this vehicle just run this service one time
                    if i == 'A2E':
                        A2E_match_state = False                 

                    final_service_list.append(i)
                    # print('SUCCESS----: ', i)
                    # print('start_service_time ------- ', start_service_time[i])
                    
                    # zero_counter = 0 直到速度为0才停止
                    while(int(row[6]) > 0):
                        row = next(r_ss)
                        counter += 1
                    stop_row.append(counter)
                    # print('success_stop_row: ', counter)
                    
                    # 异常处理--对开始时间进行最后确定
                    try:
                        final_time.append(start_service_time[i])
                    except KeyError as e:
                        final_time.append(time[0])

                    # 异常处理--对开始的行号进行记录
                    try:
                        start_row.append(start_row_1[i])
                    except KeyError as e:
                        start_row.append(start_row_2[0])  
                        
                    time = []
                    matching_service = []
                    start_service_time = {}
                    match_start_state = True
                    start_service_time_state = True
                    mid_appear_counter = 0
                    must_stop_counter = 0
                    biz2_stop_counter = 0
                    counter_start_mark = 0
                    zero_speed_counter = 0
                    LT29_counter = 0
                    A2E_match_counter = 0
                    A1_A1E_dis = []
                    start_row_1 = {}
                    start_row_2 = []
                    
                    # UT_FoS_match_state = True
                    UT_FoS_match_counter = 0
                    ## 针对每天第一趟服务的缺失假设状态进行置False，每个service只记录最早的那一次 ##
                    miss_headmost_state = False
                    miss_headmost_passby_poi = []
                    ## 针对中途丢失数据的，遇到完整的匹配，则清空该次的站点记录
                    miss_mid_passby_poi = []
                    miss_mid_time_state = True  #让下次可以记录起始时间

                    # start_POI_copy = copy.deepcopy(start_POI)
                    for j in service_copy:
                        # pointer_State[j] = True
                        pointer_POI[j] = -1
                    break

    # result_path = '/'.join(['..', 'Final_Result', 'test-' + filepath.split('_')[3]])
    # with open(result_path, 'w', newline ='') as fnew:    #use 'w' for writing str and newline='' for deleting extra idle row
    #     writer = csv.writer(fnew)
    #     for row in writer:
    #         writer.writerow()
    try:
        if final_service_list[0] == 'D1' and final_service_list[1] != 'D1' :
            final_service_list.pop(0)
            final_time.pop(0)
            start_row.pop(0)
            stop_row.pop(0)
    except IndexError as e:
        print(e)
    for i in range(len(final_service_list)):
        if final_service_list[i] != 'UT-FoS' and i > 0:
            if start_row[i]-60 <= service_time_pair[i-1][3]:
                service_time_pair.append([final_service_list[i], final_time[i], service_time_pair[i-1][3]+1, stop_row[i]+5])
            else:
                service_time_pair.append([final_service_list[i], final_time[i], start_row[i]-5, stop_row[i]+5])

        else:
            service_time_pair.append([final_service_list[i], final_time[i], start_row[i]-5, stop_row[i]+5])         
    print('result: \n', final_service_list)
    print('length: ', len(final_service_list))
    # print('service_time_pair: ')
    pprint.pprint(service_time_pair)

    dataframe = [ service_time_pair[i] for i in range(len(service_time_pair))]
    vehicle_id = filepath.split('/')[2].split('-')[0]
    columns = [ vehicle_id, 'time', 'start_row_' + vehicle_id, 'stop_row_' + vehicle_id ]
    tmp_df = pd.DataFrame(dataframe, columns = columns)
    final_time = []
    return (tmp_df, service_time_pair)

def tagService_RawData(read_filepath, write_filepath, all_record):
    counter_all = {}  #{A1:X, A2:X, .......} 记录读取了X行
    for key in all_record.keys():
        counter_all[key] = 0
    with open(read_filepath, 'r') as f:
        read = csv.reader(f)
        headings_read = next(read)
        with open(write_filepath, 'w', newline ='') as fnew:    #use 'w' for writing str and newline='' for deleting extra idle row
            writer = csv.writer(fnew)
            writer.writerow(headers)
            for row in read:
                count = 0
                write_index = 0
                write_state = 0  #0:negative 1:active
                write_start_state = 0
                write_stop_state = 0
                # counter_all[row[0]] += 1
                for count in range(len(all_record[row[0]])):
                    # 如果这辆车的行数在一个服务的起停范围内，则标注相应的服务！
                    if counter_all[row[0]] >= all_record[row[0]][count][2] and counter_all[row[0]] <= all_record[row[0]][count][3]:
                        write_state = 1                        
                        write_index = count  #需要写入服务的车在list中的index
                    if counter_all[row[0]] == all_record[row[0]][count][2]:  # 判断服务开始
                        write_start_state = 1
                        write_start = count
                        break
                    if counter_all[row[0]] == all_record[row[0]][count][3]:  #判断服务停止
                        write_stop_state = 1
                        write_stop = count
                        break
               
                counter_all[row[0]] += 1  #放在此处是为了能够将LDH的起始行时间算入

                if write_state == 1 and write_start_state == 1:  # 标志服务的起始
                    writer.writerow([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], all_record[row[0]][write_start][0], 'service_start'])
                    # print('start-- ', write_start, counter_all[row[0]])
                if write_state == 1 and write_stop_state == 1:  # 标志服务的停止
                    writer.writerow([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], all_record[row[0]][write_stop][0], 'service_stop'])
                    # print('stop-- ', write_stop, counter_all[row[0]])
                if write_state == 1 and write_start_state != 1 and write_stop_state != 1:
                    writer.writerow([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], all_record[row[0]][write_index][0], 'Running'])
                if write_state != 1:
                    writer.writerow([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], '', '', ''])
    print('Success!')

    #             if write_state == 1 and write_start_state == 1:  # 标志服务的起始
    #                 writer.writerow([row[0], row[1], row[2], row[8], all_record[row[0]][write_index][0], 'service_start'])
    #             elif write_state == 1 and write_stop_state == 1:  # 标志服务的停止
    #                 writer.writerow([row[0], row[1], row[2], row[8], all_record[row[0]][write_index][0], 'service_stop'])
    #             elif write_state == 1:
    #                 writer.writerow([row[0], row[1], row[2], row[8], all_record[row[0]][write_index][0]])
    #             else:
    #                 writer.writerow([row[0], row[1], row[2], 'Running'])

    # print('Success!')

def run(folderpath, read_filepath, write_filepath):
    '''Aggregate all functions and run this module
    
    folderpath: VisitAllFile(folderpath)
    '''
    all_poi_normal_num = {}
    all_poi_normal = getStations_Inorder()
    for i in service:
        all_poi_normal_num[i] = len(all_poi_normal[i])
    # print('all_poi_normal_num: ', all_poi_normal_num)

    df_row = pd.DataFrame()
    all_record = {}
    for filePath in VisitAllFile(folderpath)[0]:
        dir_filePath = '/'.join([folderpath, filePath])
        tmp_df = match_Service(dir_filePath, all_poi_normal, all_poi_normal_num)[0]
        df_row = pd.concat([df_row, tmp_df], axis = 1)
        date = '-'.join(filePath.split('-')[1:])
        result_path = ''.join(['Final_Result/', 'test-dfrow-', date])
        MyFunctions.checkPath(result_path)
        df_row.to_csv(result_path)

        service_time_pair = match_Service(dir_filePath, all_poi_normal, all_poi_normal_num)[1]
        all_record[filePath.split('-')[0]] = service_time_pair
    # pprint.pprint(all_record)   #{2023:[BTC1, start_row, stop_row, BTC2, start_row, stop_row, ....], 2024:...}
    
    #Last step to write final result on new_raw data
    tagService_RawData(read_filepath, write_filepath, all_record)


############## main ###############
if __name__ == "__main__":
    folderpath = 'fenced_vehicle/2017-09-05'
    read_filepath = 'fenced/fenced-2017-09-05.csv'
    write_filepath = 'Final_Result/final_2017-09-05.csv'
    run(folderpath, read_filepath, write_filepath)
    



    # print('---------------------------------- main -----------------------------------')
    # all_poi_normal_num = {}
    # all_poi_normal = getStations_Inorder()
    # for i in service:
    #     all_poi_normal_num[i] = len(all_poi_normal[i])
    # print('all_poi_normal_num: ', all_poi_normal_num)

    # df_row = pd.DataFrame()
    # all_record = {}
    # for filePath in VisitAllFile('fenced_vehicle')[0]:
    #     dir_filePath = '/'.join(['fenced_vehicle', filePath])

    #     tmp_df = match_Service(dir_filePath, all_poi_normal, all_poi_normal_num)[0]

    #     df_row = pd.concat([df_row, tmp_df], axis = 1)
    #     result_path = '/'.join(['Final_Result', 'test-dfrow-' + filePath.split('_')[1]])
    #     df_row.to_csv(result_path)

    #     service_time_pair = match_Service(dir_filePath, all_poi_normal, all_poi_normal_num)[1]
    #     all_record[filePath.split('_')[0]] = service_time_pair
    # print(all_record)   #{2023:[BTC1, [start_row, stop_row], BTC2, [start_row, stop_row], ....], 2024:...}

    # read_filepath = 'fenced/fenced2018-01-08.csv'
    # write_filepath = 'Final_Result/final_22018-01-08.csv'
    # tagService_RawData(read_filepath, write_filepath, all_record)
        








