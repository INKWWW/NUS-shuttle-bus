#!/usr/bin/python
# -*-coding: UTF-8-*-
# @Author :Wang Hanmo
# 
'''Detect service and tag them'''
from openpyxl import load_workbook   #pip install openpyxl
from collections import namedtuple
import pandas as pd
import csv
import MyFunctions
import math
import copy
import pprint
import os
import pdb  #breakpoint debug

start_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'BIZ2', 'Botanic Gardens MRT', 'Ventus (Opp LT13)', 'University Town']
# stop_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'Car Park 11', 'BTC - Oei Tiong Ham Building']
headers = ['node_id','vehicle_serial','gps_time','latitude', 'longitude', 'altitude', 'speed', 'heading', 'POI', 'service', 'service_start_stop']
service = ['A1', 'A2', 'B1', 'B2', 'C*', 'C', 'D1', 'D2', 'A1E', 'A2E', 'BTC1', 'BTC2', 'UT-FoS'] #13
start_POI = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'BIZ2', 'Opp HSSML', 'Botanic Gardens MRT', 'University Town']
must_stop_point = ['Prince George\'s Park Terminal', 'Kent Ridge Bus Terminal', 'Botanic Gardens MRT']
all_polygon = {}
# read_file_name = '../fenced_vehicle/2024_inside_2017-09-04.csv'


def VisitAllFile(folderpath):
    '''Visit all files in a folder and get these files' names
    '''
    fileName = [filenames for (dirpath, dirnames, filenames) in os.walk(folderpath)]
    # print('FILENAME: ', fileName)
    return fileName

def get_NUS_fence_vertex():
    '''Get all the vertexs on the NUS_fence orderly
    '''
    wpoly = load_workbook('../NUS_polygon.xlsx')
    sheet = wpoly.get_sheet_names()
    poly = wpoly.get_sheet_by_name(sheet[0])
    all_rows = poly.max_row
    all_columns = poly.max_column
    # print('max_polygon_columns : ', all_columns)
    # print('max_polygon_rows : ', all_rows)
    polygon = []
    coordinate = ()
    j = 0
    for col in range (4,39,3):
        j += 1
        polygon = []
        for i in range(3, all_rows + 1) :
            coordinate = (poly.cell(row = i, column = col).value, poly.cell(row = i, column = col+1).value)
            # print(coordinate)
            if coordinate[0] is not None:
                polygon.append(coordinate)
        all_polygon[service[j-1]] = polygon

def getStations_Inorder():
    '''Get the stations in order of every service--this is used to match
    '''
    poi_read = load_workbook('../verify_service_route_6.xlsx')
    poi_sheet = poi_read.get_sheet_names()
    poi = poi_read.get_sheet_by_name(poi_sheet[0])
    poi_all_row = poi.max_row
    poi_all_column = poi.max_column
    # print('--------------------------POI:----------------------------')
    # print('max_poi_rows: ', poi_all_row)
    # print('max_poi_columns: ', poi_all_column)
    # print(poi.cell(row = 15, column = 1).value)
    # print(type(poi.cell(row = 15, column = 1).value))
    # pdb.set_trace()
    all_poi_normal = {}
    poi_normal = []
    poi_order = ''
    all_poi_normal_num = {}
    j = 0
    for poi_col in range (1, 14):
        j += 1
        poi_normal = []
        for i in range (2, poi_all_row + 1):
            poi_order = poi.cell(row = i, column = poi_col).value
            # print(poi_order)
            if poi_order is not None:
                poi_normal.append(poi_order)
            else:
                continue
        all_poi_normal[service[j-1]] = poi_normal
        all_poi_normal_num[service[j-1]] = len(poi_normal)
    # pprint.pprint(all_poi_normal)
    # print(all_poi_normal_num)
    return(all_poi_normal)
    # print('------')
    # print(all_poi_normal.items())
    # pdb.set_trace()

def getStations_All():
    '''Get the stations in order of every service--this is the standard / original one
    '''
    poi_read = load_workbook('../verify_service_route_all.xlsx')
    poi_sheet = poi_read.get_sheet_names()
    poi = poi_read.get_sheet_by_name(poi_sheet[0])
    poi_all_row = poi.max_row
    poi_all_column = poi.max_column
    
    all_poi = {}
    poi_normal = []
    poi_order = ''
    all_poi_num = {}
    j = 0
    for poi_col in range (1, 14):
        j += 1
        poi_normal = []
        for i in range (2, poi_all_row + 1):
            poi_order = poi.cell(row = i, column = poi_col).value
            # print(poi_order)
            if poi_order is not None:
                poi_normal.append(poi_order)
            else:
                continue
        all_poi[service[j-1]] = poi_normal
        all_poi_num[service[j-1]] = len(poi_normal)
    print('all_poi')
    return(all_poi)

def match_Service(filepath, all_poi_normal, all_poi_normal_num):
    '''Match the Service according to the order of POIs in different service
    '''
    service_copy = copy.deepcopy(service)
    start_POI_copy = copy.deepcopy(start_POI)
    final_service_list = []
    matching_service = []
    A1_A1E_dis = []
    C_CKV_dis =[]
    time = []
    final_time = []
    start_row = []
    stop_row = []
    service_time_pair = []
    start_service_time = {}
    start_row_1 = {}
    start_row_2 = []
    LT29_counter = 0
    zero_speed_counter = 0
    counter = 0
    mid_appear_counter = 0
    start_match = 0
    counter_start_mark = 0
    A2E_match_counter = 0
    must_stop_counter = 0
    biz2_stop_counter = 0
    LT29_counter_state = True
    # particular_must_stop_poi = []
    match_start_state = True
    A2E_match_state = True
    check = False
    start_service_time_state = True

    UT_FoS_hour_list = [9, 11, 13, 15]
    UT_FoS_match_counter = 0
    UTFOS_time_hour = 0
    UT_FoS_match_state = True

    pointer_State = {}     # record_start:True
    for i in service_copy:
        pointer_State[i] = True
    # print(pointer_State)

    pointer_POI = {}
    for i in service_copy:
        pointer_POI[i] = -1
    # print(pointer_POI)
    # pdb.set_trace()

    with open(filepath, 'r') as ss_v:  #with打开不用在意close
        r_ss = csv.reader(ss_v)
        for row in r_ss:
            counter += 1
            # print(counter)
            ## Define the start of the process of matching
            
            # Because there are only two A2E in the afternoon (about 17:30 & 17:40) - different vehicles runs these two service 
            # what's more, 不会让跑A1 或者 A2的车去跑
            if row[8] == 'Ventus (Opp LT13)' and A2E_match_state:
                A2E_match_counter += 1
                if A2E_match_counter > 12:                    
                    time_hour = int(row[2].split('T')[1].split(':')[0].strip())  #strip()消除字符前后可能出现的空格
                    time_min = int(row[2].split('T')[1].split(':')[1].strip())
                    if time_hour == 17 and ('A1' not in matching_service) and ('A2' not in matching_service):
                        pointer_POI['A2E'] = 0
                        A2E_match_counter = 0
                        matching_service.append('A2E')
                        start_service_time['A2E'] = row[2]
                        start_row_1['A2E'] = counter
                        if len(matching_service) > 0:
                            print('@@@ A2E @@@ ', matching_service)
                        # print('counter:  ', counter)
                        # print('time: ', row[0])
                        # time.append(row[2])
            else:
                A2E_match_counter = 0

            # 决定UT-FoS是否需要加入到matching_list里面
            if (len(final_service_list) > 0) and ('D2' not in matching_service) and ('D2' not in final_service_list):
                UT_FoS_match_state = True
            if row[8] == 'University Town' and int(row[6]) == 0 and UT_FoS_match_state:
                UT_FoS_match_counter += 1
                if UT_FoS_match_counter >= 10:
                    UTFOS_time_hour = int(row[2].split('T')[1].split(':')[0].strip())  #strip()消除字符前后可能出现的空格
                    if UTFOS_time_hour in UT_FoS_hour_list:
                        pointer_POI['UT-FoS'] = 0
                        UT_FoS_match_counter = 0
                        matching_service.append('UT-FoS')
                        UT_FoS_hour_list.pop(UT_FoS_hour_list.index(UTFOS_time_hour))
                        start_service_time['UT-FoS'] = row[2]
                        start_row_1['UT-FoS'] = counter
                        UT_FoS_match_state = False
                        # if len(matching_service) > 0:
                        #     print('@@@ UT-FoS @@@ ', matching_service)
                        #     print('#$@#@%',UT_FoS_hour_list)               


            # others except A2E
            if match_start_state and row[8] in start_POI_copy:
                counter_start_mark += 1
            else:
                counter_start_mark = 0

            if match_start_state and counter_start_mark > 0:             
                for service_item in service_copy:              
                    if (all_poi_normal[service_item][0] == row[8]) and (row[8] in start_POI_copy) and service_item != 'UT-FoS':
                        pointer_POI[service_item] = 0
                        # pointer_State[service_item] = False
                    # 匹配到最后一个'UT-FoS'为止
                    if service_item == 'UT-FoS' and (0 in [values for key, values in pointer_POI.items()]):
                        match_start_state = False
                matching_service = [key for key, values in pointer_POI.items() if values == 0]  #such as [A1, A2, B1]
                # if len(matching_service) > 0:
                #     print('@@', matching_service)
                # print('counter: ----- ', counter)
                # print('time:', row[0])            

            #D2  VS  UT-FoS  // D2 pass 2 times of LT29
            if row[8] != 'LT29':
                LT29_counter_state = True
            if row[8] == 'LT29' and LT29_counter_state:
                LT29_counter += 1
                LT29_counter_state = False

            ## match
            for i in matching_service:

                # continuously, velocity = 0 for a long time --> re_match
                if int(row[6]) <= 1 and row[8] != all_poi_normal[i][0]:
                    zero_speed_counter += 1
                if int(row[6]) > 1:
                    zero_speed_counter = 0

                # A1  VS  A1E
                if ('A1' in matching_service) and ('A1E' in matching_service):
                    if row[8] not in A1_A1E_dis or row[8] != A1_A1E_dis[len(A1_A1E_dis) - 1]:
                        A1_A1E_dis.append(row[8])

                if pointer_POI[i] == 1 and int(row[6]) == 0 and start_service_time_state and (i != 'UT-FoS'):
                    start_service_time[i] = row[2]
                    # print("!!!!!start !!!!!!!", start_service_time)
                    start_row_1[i] = counter

                ## Match service from the first POI in a fixed order
                PGP_Hse_list = ['PGP Hse No. 7', 'PGP Hse No. 12', 'PGP Hse No. 14 and 15']
                #第一行：针对普通正常无缺失的匹配
                #第二行：针对PGP Hse No.处的缺失进行匹配
                #第三行：针对UHall附近的缺失进行匹配
                if row[8] == all_poi_normal[i][pointer_POI[i]] or \
                    ((row[8] in PGP_Hse_list) and (all_poi_normal[i][pointer_POI[i]] in PGP_Hse_list)) or \
                    ((row[8] == 'Opp UHall') and (all_poi_normal[i][pointer_POI[i]] == 'UHall')):
                    # print(row)
                    pointer_POI[i] += 1
                    
                    if pointer_POI[i] == 2:
                        # print('start_service_time: ', row[2])
                        # print('start_row : ', counter)
                        if row[2] not in time:
                            time.append(row[2])
                            start_row_2.append(counter)
                    
                    # # A1  VS  A1E
                    # if ('A1' in matching_service) and ('A1E' in matching_service):
                    #     if row[8] not in A1_A1E_dis:
                    #         A1_A1E_dis.append(row[8])

                    # C  VS  C*
                    if ('C' in matching_service) and ('C*' in matching_service):
                        if row[8] not in C_CKV_dis:
                            C_CKV_dis.append(row[8])

                        # if row[8] not in A1_A1E_dis:
                        #     A1_A1E_dis.append(row[8])
                            # print('A1_A1E_dis: ', A1_A1E_dis)
                    # pdb.set_trace()
                    
                if (row[8] != all_poi_normal[i][0]):    ##!!!!!improvement
                    check = True
                    # print('!!!',counter)
                
                # prevent the ahead apperance of first station  --> quit and restart
                if pointer_POI[i] > 0 and check:
                    if row[8] == all_poi_normal[i][0] and int(row[6]) == 0:
                        mid_appear_counter += 1   #在for循环里面，每行都会自加len(matching_service)次
                        # print(row)
                        # print('-?-', mid_appear_counter)
                    else:
                        mid_appear_counter = 0
                        
                    # must stop and restart matching
                    if row[8] in must_stop_point and ('BTC1' not in matching_service) and ('D2' not in matching_service):
                        must_stop_counter += 1    #在for循环里面，每行都会自加len(matching_service)次
                    
                        # must_stop_counter = 0
                        # print('must stop poi:', must_stop_counter)
                        # print('counter is ', counter)
                    if ('B2' in matching_service or 'D1' in matching_service) and row[8] == 'BIZ2' and int(row[6]) <= 1:
                        biz2_stop_counter += 1
                        # print('biz2_stop_counter: ',biz2_stop_counter)
                    if row[8] != 'BIZ2':
                        biz2_stop_counter = 0

                    if (mid_appear_counter // len(matching_service) >= 35) or \
                        (must_stop_counter // len(matching_service) >= 36) or \
                        (biz2_stop_counter // len(matching_service) >= 50) or \
                        (zero_speed_counter // len(matching_service) > 165):
                        # print('mid_appear_counter -- :', mid_appear_counter // len(matching_service))
                        # print('must_stop_counter -- :', must_stop_counter // len(matching_service))
                        # print('biz2_stop_counter -- :', biz2_stop_counter // len(matching_service))
                        # print('zero_speed_counter -- :', zero_speed_counter // len(matching_service))

                        match_start_state = True
                        check = False
                        start_service_time_state = True
                        mid_appear_counter = 0
                        must_stop_counter = 0
                        biz2_stop_counter = 0
                        counter_start_mark = 0
                        zero_speed_counter = 0
                        LT29_counter = 0
                        A2E_match_counter = 0
                        matching_service = []
                        A1_A1E_dis = []
                        start_service_time = {}
                        # UT_FoS_match_state = True
                        UT_FoS_match_counter = 0
                    
                        print('pause counter: -- ', counter)
                        
                        time = []
                        # if pointer_POI[i] < all_poi_normal_num[i]:
                        #     start_POI_copy.pop(start_POI_copy.index(all_poi_normal[i][0]))
                        for j in service_copy:
                            # pointer_State[j] = True
                            pointer_POI[j] = -1
                        break                       

                # Have found the final service!! Then quit, reset all parameters and restart this program loop
                if pointer_POI[i] == all_poi_normal_num[i]:
                    # A1  VS A1E
                    # if (i == 'A1E') and ('COM 2 (CP13)' in A1_A1E_dis):
                    # print('A1_A1E_dis: ', A1_A1E_dis)
                    
                    if i == 'A1E' and A1_A1E_dis[A1_A1E_dis.index('COM 2 (CP13)') + 1] == 'Opp NUSS':
                        # print('A1_A1E_dis: ', A1_A1E_dis)
                        # matching_service.pop(matching_service.index('A1E'))
                        # print('pop :', matching_service)
                        i = 'A1'
                        A1_A1E_dis = []
                        # break
                    
                    try:
                        if i == 'A1' and A1_A1E_dis[A1_A1E_dis.index('COM 2 (CP13)') + 1] == 'Opp HSSML':
                            i = 'A1E'
                            A1_A1E_dis = []
                    except ValueError as e:
                        i = 'A1E'
                        A1_A1E_dis = []


                    # D2  VS  UT-FoS
                    # scenario 1: 假如被判断成UT-FoS, 但是不在应该运行的时间内，则判断成'D2-S‘--特殊的D2，可能是确实了一些数据的
                    UT_FoS_hour = 0  
                    if (i == 'UT-FoS') and 'D2' in matching_service:
                        UT_FoS_hour = int(row[2].split('T')[1].split(':')[0].strip())
                        if UT_FoS_hour not in [9, 11, 13, 15]:
                            i = 'D2-S'

                    # scenatio 2: 假如被判断成D2，则判断经过了几个LT29，以此来判断是否为UT-FoS
                    if (i == 'D2') and LT29_counter > 3 and ('D2' not in final_service_list):
                        UT_FoS_hour = int(row[2].split('T')[1].split(':')[0].strip())
                        if UT_FoS_hour in UT_FoS_hour_list:
                            i = 'UT-FoS'
                        # print('LT29_counter:', LT29_counter)

                    # C  VS  C*
                    if i == 'C*' and 'Computer Centre' in C_CKV_dis:
                        matching_service.pop(matching_service.index('C*'))
                        C_CKV_dis = []
                        break

                    # Let this vehicle just run this service one time
                    if i == 'A2E':
                        A2E_match_state = False                 

                    final_service_list.append(i)
                    print('SUCCESS----: ', i)
                    # print('start_service_time ------- ', start_service_time[i])
                    print('success_stop_row:  ', counter)
                    stop_row.append(counter)
                    
                    # 异常处理--对开始时间进行最后记录
                    try:
                        final_time.append(start_service_time[i])
                    except KeyError as e:
                        final_time.append(time[0])

                    # 异常处理--对开始的行号进行记录
                    try:
                        start_row.append(start_row_1[i])
                    except KeyError as e:
                        start_row.append(start_row_2[0])  
                        
                    time = []
                    matching_service = []
                    start_service_time = {}
                    match_start_state = True
                    start_service_time_state = True
                    mid_appear_counter = 0
                    must_stop_counter = 0
                    biz2_stop_counter = 0
                    counter_start_mark = 0
                    zero_speed_counter = 0
                    LT29_counter = 0
                    A2E_match_counter = 0
                    A1_A1E_dis = []
                    start_row_1 = {}
                    start_row_2 = []
                    start_service_time_state = True
                    # UT_FoS_match_state = True
                    UT_FoS_match_counter = 0

                    # start_POI_copy = copy.deepcopy(start_POI)
                    for j in service_copy:
                        # pointer_State[j] = True
                        pointer_POI[j] = -1
                    break
    # result_path = '/'.join(['..', 'Final_Result', 'test-' + filepath.split('_')[3]])
    # with open(result_path, 'w', newline ='') as fnew:    #use 'w' for writing str and newline='' for deleting extra idle row
    #     writer = csv.writer(fnew)
    #     for row in writer:
    #         writer.writerow()

    for i in range(len(final_service_list)):
        if final_service_list[i] != 'UT-FoS':
            service_time_pair.append([final_service_list[i], final_time[i], start_row[i]-100, stop_row[i]+100])         
        else:
            service_time_pair.append([final_service_list[i], final_time[i], start_row[i], stop_row[i]])      
    print('result: \n', final_service_list)
    print('length: ', len(final_service_list))
    # pprint.pprint(service_time_pair)

    dataframe = [ service_time_pair[i] for i in range(len(service_time_pair))]
    vehicle_id = filepath.split('/')[2].split('-')[0]
    columns = [ vehicle_id, 'time', 'start_row_' + vehicle_id, 'stop_row_' + vehicle_id ]
    tmp_df = pd.DataFrame(dataframe, columns = columns)
    final_time = []
    return (tmp_df, service_time_pair)

def tagService_RawData(read_filepath, write_filepath, all_record):
    counter_all = {}  #{A1:X, A2:X, .......} 记录读取了X行
    for key in all_record.keys():
        counter_all[key] = 0
    with open(read_filepath, 'r') as f:
        read = csv.reader(f)
        headings_read = next(read)
        with open(write_filepath, 'w', newline ='') as fnew:    #use 'w' for writing str and newline='' for deleting extra idle row
            writer = csv.writer(fnew)
            writer.writerow(headers)
            for row in read:
                count = 0
                write_index = 0
                write_state = 0  #0:negative 1:active
                write_start_state = 0
                write_stop_state = 0
                counter_all[row[0]] += 1
                for count in range(len(all_record[row[0]])):
                    # 如果这辆车的行数在一个服务的起停范围内，则标注相应的服务！
                    if counter_all[row[0]] >= all_record[row[0]][count][2] and counter_all[row[0]] <= all_record[row[0]][count][3]:
                        write_state = 1                        
                        write_index = count  #需要写入服务的车在list中的index
                    if counter_all[row[0]] == all_record[row[0]][count][2]:  # 判断服务开始
                        write_start_state = 1
                    if counter_all[row[0]] == all_record[row[0]][count][3]:  #判断服务停止
                        write_stop_state = 1

                if write_state == 1 and write_start_state == 1:  # 标志服务的起始
                    writer.writerow([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], all_record[row[0]][write_index][0], 'service_start'])
                elif write_state == 1 and write_stop_state == 1:  # 标志服务的停止
                    writer.writerow([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], all_record[row[0]][write_index][0], 'service_stop'])
                elif write_state == 1:
                    writer.writerow([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], all_record[row[0]][write_index][0]])
                else:
                    writer.writerow([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], 'Running'])
    print('Success!')

    #             if write_state == 1 and write_start_state == 1:  # 标志服务的起始
    #                 writer.writerow([row[0], row[1], row[2], row[8], all_record[row[0]][write_index][0], 'service_start'])
    #             elif write_state == 1 and write_stop_state == 1:  # 标志服务的停止
    #                 writer.writerow([row[0], row[1], row[2], row[8], all_record[row[0]][write_index][0], 'service_stop'])
    #             elif write_state == 1:
    #                 writer.writerow([row[0], row[1], row[2], row[8], all_record[row[0]][write_index][0]])
    #             else:
    #                 writer.writerow([row[0], row[1], row[2], 'Running'])

    # print('Success!')

def run(folderpath, read_filepath, write_filepath):
    '''Aggregate all functions and run this module
    
    folderpath: VisitAllFile(folderpath)
    '''
    all_poi_normal_num = {}
    all_poi_normal = getStations_Inorder()
    for i in service:
        all_poi_normal_num[i] = len(all_poi_normal[i])
    # print('all_poi_normal_num: ', all_poi_normal_num)

    df_row = pd.DataFrame()
    all_record = {}
    for filePath in VisitAllFile(folderpath)[0]:
        dir_filePath = '/'.join([folderpath, filePath])
        tmp_df = match_Service(dir_filePath, all_poi_normal, all_poi_normal_num)[0]
        df_row = pd.concat([df_row, tmp_df], axis = 1)
        date = '-'.join(filePath.split('-')[1:])
        result_path = ''.join(['Final_Result/', 'test-dfrow-', date])
        MyFunctions.checkPath(result_path)
        df_row.to_csv(result_path)

        service_time_pair = match_Service(dir_filePath, all_poi_normal, all_poi_normal_num)[1]
        all_record[filePath.split('-')[0]] = service_time_pair
    # print(all_record)   #{2023:[BTC1, [start_row, stop_row], BTC2, [start_row, stop_row], ....], 2024:...}
    
    #Last step to write final result on new_raw data
    tagService_RawData(read_filepath, write_filepath, all_record)


############## main ###############
if __name__ == "__main__":
    folderpath = 'fenced_vehicle/2017-09-05'
    read_filepath = 'fenced/fenced-2017-09-05.csv'
    write_filepath = 'Final_Result/final_2017-09-05.csv'
    run(folderpath, read_filepath, write_filepath)
    



    # print('---------------------------------- main -----------------------------------')
    # all_poi_normal_num = {}
    # all_poi_normal = getStations_Inorder()
    # for i in service:
    #     all_poi_normal_num[i] = len(all_poi_normal[i])
    # print('all_poi_normal_num: ', all_poi_normal_num)

    # df_row = pd.DataFrame()
    # all_record = {}
    # for filePath in VisitAllFile('fenced_vehicle')[0]:
    #     dir_filePath = '/'.join(['fenced_vehicle', filePath])

    #     tmp_df = match_Service(dir_filePath, all_poi_normal, all_poi_normal_num)[0]

    #     df_row = pd.concat([df_row, tmp_df], axis = 1)
    #     result_path = '/'.join(['Final_Result', 'test-dfrow-' + filePath.split('_')[1]])
    #     df_row.to_csv(result_path)

    #     service_time_pair = match_Service(dir_filePath, all_poi_normal, all_poi_normal_num)[1]
    #     all_record[filePath.split('_')[0]] = service_time_pair
    # print(all_record)   #{2023:[BTC1, [start_row, stop_row], BTC2, [start_row, stop_row], ....], 2024:...}

    # read_filepath = 'fenced/fenced2018-01-08.csv'
    # write_filepath = 'Final_Result/final_22018-01-08.csv'
    # tagService_RawData(read_filepath, write_filepath, all_record)
        








