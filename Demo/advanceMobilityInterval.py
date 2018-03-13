#!/usr/bin/python
# -*-coding: UTF-8-*-
# @Author :Wang Hanmo
# 
'''Detect service and tag them'''
from openpyxl import load_workbook   #pip install openpyxl
from collections import namedtuple
import csv
import MyFunctions
import pprint
from decimal import Decimal
import pdb  #breakpoint debug

service = ['A1', 'A2', 'B1', 'B2', 'C*', 'C', 'D1', 'D2', 'A1E', 'A2E', 'BTC1', 'BTC2', 'UT-FoS']  #13
headers = ['node_id', 'vehicle_serial', 'date', 'time', 'POI', 'service', 'interval']

service_times = {}  #{service: times}
service_ss = {}  #{service: [[start_time, stop_time], [start_time, stop_time]]}
# service_interval = {}  #{service: [interval1, interval2, ...]}
POI_services_interval = {}  # {YIH: [{A1:[T1, T2, T3]}, {A2:[T4, T5, T6]}, {B1:[T7, T8, T9]}]}
POI_services_start = {}  # {YIH: [{A1:[S1, S2, S3]}, {A2:[S4, S5, S6]}, {B1:[S7, S8, S9]}]}
start_row_num = {}  # {YIH: [{A1:[row1, row2, row3]}, {A2:[row1, row2, row3]}, {B1:[row1, row2, row3]}]}

'''Get the stations in order of every service'''
def getAllPOI():
    poi_read = load_workbook('../verify_service_route_7.xlsx')
    poi_sheet = poi_read.get_sheet_names()
    poi = poi_read.get_sheet_by_name(poi_sheet[0])
    poi_all_row = poi.max_row
    poi_all_column = poi.max_column
    all_poi_normal = {}
    poi_normal = []
    poi_order = ''
    all_poi_normal_num = {}
    all_POI = []
    POI_services = {}  # {YIH: [A1, A2 B1], UHall: [D1, A1]}
    j = 0
    for poi_col in range (1, 14):
        j += 1
        poi_normal = []
        for i in range (2, poi_all_row + 1):
            poi_order = poi.cell(row = i, column = poi_col).value
            if poi_order is not None:
                poi_normal.append(poi_order)
            else:
                continue
        all_poi_normal[service[j-1]] = poi_normal   ##{'A1':[XX, XX, XX, XX], 'B2':[XX, XX, XX, XX]...}
        all_poi_normal_num[service[j-1]] = len(poi_normal)
        # pprint.pprint(all_poi_normal)

    ## 统计所有的站点POI
    for key in all_poi_normal.keys():
        for m in all_poi_normal[key]:
            if m not in all_POI:
                all_POI.append(m)
    # pprint.pprint(all_POI)

    ## 统计所有POI对应的service
    for poi in all_POI:
        POI_services[poi] = []
        for key in all_poi_normal.keys():
            if poi in all_poi_normal[key]:
                POI_services[poi].append(key)
    # pprint.pprint(POI_services)
    return(all_poi_normal, all_POI, POI_services)


def recordStartStop(read_filepath, poi, service):
    '''Fix POI, then find relative service one by one and calculate intervals. Next, go to next POI
    
    [description]: 1. Fix POI  2. Fix service  3. Find the service_ss = {}  #{service: [start_time, stop_time]} according to the change of 'node_id'

    [Attention]: 'service' from service column should be just extracted the first two letter!
    
    Arguments:
        read_filepath {[str]} -- [read from this path]
        poi : str -- POI
        service : str -- service
        # write_filepath {[str]} -- [write into this path]
    '''
    start_time_state = True
    stop_time_state = False
    # start_state_counter 0  # 计数2个后才判断该车进入，则作为离站时间进行记录
    stop_state_counter = 0  # 计数5个后如果还没该车出现，则作为离站时间进行记录
    counter_start = 0  # 记录每个poi开始的行数，方便你后面写文件快速

    service_times = {}  #{service: times}
    service_ss = {}  #{service: [[start_time, stop_time], [start_time, stop_time]...]} -- [进站时间，出站时间]
    service_interval = {}  #{service: [interval1, interval2, ...]}
    start_stop = []
    start_row_service = {}  # {A1:[row1, row2, row3]} --> a dict in start_row_row[service]

    if poi not in POI_services_interval.keys():
        POI_services_interval[poi] = []

    with open(read_filepath, 'r') as f:
        service_interval[service] = []
        service_ss[service] = []
        start_row_service[service] = []
        former_nodeid = ''
        r_f = csv.reader(f)
        headline = next(r_f)
        for row in r_f:
            counter_start += 1
            if len(row) == 11: 
                if row[8] == poi:
                    # service进入站点
                    if row[9] == service and start_time_state:
                        date = row[2].split('T')[0]
                        time = row[2].split('T')[1].split('.')[0]
                        # 记录达到站点的时间                    
                        start_stop.append(time)
                        ## !! ##
                        # 会记录下最后一次服务的开始行数，但是最后一次服务没有duration了，所以后面writeIntervalInfo的时候，
                        # except error后需要提示为最后一次服务，然后继续运行程序
                        start_row_service[service].append(counter_start)  
                        start_time_state = False
                        continue

                    # 数5行后续的站点，作为判断真的离开的标志
                    if (not start_time_state) and row[9] != service:
                        stop_state_counter += 1
                    else:
                        stop_state_counter = 0
                        # 随时记录时间,确保结束时间的记录准确
                        date = row[2].split('T')[0]
                        time = row[2].split('T')[1].split('.')[0]             
                  
                    if stop_state_counter > 5:
                        stop_time_state = True
                        stop_state_counter = 0

                    # service离开该站点  
                    if stop_time_state:
                        start_time_state = True
                        stop_time_state = False
                        # date = row[2].split('T')[0]
                        # time = row[2].split('T')[1].split('.')[0]
                        start_stop.append(time)

                        service_ss[service].append(start_stop)
                        # print(service_ss)
                        start_stop = []
                        # print(service_ss)
            else:
                continue
        
    # print(service_ss)
    return(service_ss, start_row_service)
    print('Recording the start and stop time...')


def calculateDuration(poi, service, service_ss, POI_services_interval):
    '''calculate the duration according to start_time and stop_time
    
    Arguments:
        start_time {[str]} -- [description]
        stop_time {[str]} -- [description]
        POI_services_interval = {}  # {YIH: [{A1:[T1, T2, T3]}, {A2:[T4, T5, T6]}, {B1:[T7, T8, T9]}]}
        service_interval = {}  # {service: [interval1, interval2, ...]}
    '''
    service_interval = {}
    service_interval[service] = []
    index = 0  # index in the service_ss value list
    while(index < len(service_ss[service]) - 1):
        last_leave_time = service_ss[service][index][1]
        # print(last_leave_time)
        next_arrive_time = service_ss[service][index+1][0]
        # print(next_arrive_time)
        interval = MyFunctions.calculateDuration(last_leave_time, next_arrive_time)
        service_interval[service].append(interval)
        index += 1
    POI_services_interval[poi].append(service_interval)
    # print(POI_services_interval)
    print('Successfully generate POI_services_interval')
    return(POI_services_interval)


def writeIntervalInfo(read_filepath, write_filepath, POI_services_interval, start_row_num):
    '''After getting all the infomation about Intervals, generate a file to display these infomation
    
    Arguments:
        read_filepath {[str]} -- [description]
        write_filepath {[str]} -- [description]
        POI_services_interval = {}  # {YIH: [{A1:[T1, T2, T3]}, {A2:[T4, T5, T6]}, {B1:[T7, T8, T9]}]}
        start_row_num = {}  # {YIH: [{A1:[row1, row2, row3]}, {A2:[row1, row2, row3]}, {B1:[row1, row2, row3]}]}
    
    Contents to write: headers = ['node_id', 'vehicle_serial', 'date', 'time', 'POI', 'service', 'interval']
    ''' 
    # w_fnew.writerow(row[0], row[1], date, row[2], row[8], row[9], service_start_time[service][index_poi_service])              
    with open(read_filepath, 'r') as f:
        r_f = csv.reader(f)
        headline = next(r_f)
        service_index = -1
        counter_row = 0
        write_index = 0
        with open(write_filepath, 'w', newline ='') as fnew:    #use 'w' for writing str and newline='' for deleting extra idle row
            w_fnew = csv.writer(fnew)
            w_fnew.writerow(headers)
            for row in r_f:
                counter_row += 1
                service_index = -1
                if len(row) == 11:
                    for item in start_row_num[row[8]]:
                        service_index += 1
                        if row[9] == list(item.keys())[0]:
                            if counter_row in start_row_num[row[8]][service_index][row[9]]:
                                date = row[2].split('T')[0]
                                write_index = start_row_num[row[8]][service_index][row[9]].index(counter_row)
                                try:
                                    w_fnew.writerow([row[0], row[1], date, row[2], row[8], row[9], POI_services_interval[row[8]][service_index][row[9]][write_index]])
                                except IndexError as e:
                                    print(e)
                                    print('This is the last round for this service! No more duration for this round!')
                                break            
                                                 
    print('Successfully generate new file')
                           

if __name__ == '__main__':

    all_poi_normal = getAllPOI()[0]
    all_POI = getAllPOI()[1]
    POI_services = getAllPOI()[2]
    pprint.pprint(POI_services)
    # pdb.set_trace()

    read_filepath = 'Final_Result/final_2017-09-05.csv'

    # Get all info about POI_services_interval = {}  # {YIH: [{A1:[T1, T2, T3]}, {A2:[T4, T5, T6]}, {B1:[T7, T8, T9]}]}
    # POI_services_start = {}  # {YIH: [{A1:[S1, S2, S3]}, {A2:[S4, S5, S6]}, {B1:[S7, S8, S9]}]}

    for poi in all_POI:
        print('POI------: ', poi)
        POI_services_interval[poi] = []  # Initialize for different poi
        POI_services_start[poi] = []
        start_row_num[poi] = []  # {YIH: [{A1:[row1, row2, row3]}, {A2:[row1, row2, row3]}, {B1:[row1, row2, row3]}]}
        for service in POI_services[poi]:
            # use 'start_list'(dict) to record the start of every service for a particular poi
            start_list = {}
            start_list[service] = []
            service_ss = recordStartStop(read_filepath, poi, service)[0]
            start_row_num[poi].append(recordStartStop(read_filepath, poi, service)[1])
            # print(start_row_num)
            # for item in service_ss[service]:
            #     start_list[service].append(item[0])
            # POI_services_start[poi].append(start_list)
            # print(POI_services_start)
            # pdb.set_trace()
            final_POI_services_interval = calculateDuration(poi, service, service_ss, POI_services_interval)

    # write the interval information into a file
    newpath = 'interval_' + read_filepath.split('/')[1].split('_')[1]
    write_filepath = '/'.join(['Interval', newpath])
    print(write_filepath)
    MyFunctions.checkPath(write_filepath)
    print('writing... Please wait')
    writeIntervalInfo(read_filepath, write_filepath, final_POI_services_interval, start_row_num)

